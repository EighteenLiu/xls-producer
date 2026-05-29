# -*- coding: utf-8 -*-
"""
Generate the monthly street problem summary workbook.

The source ledger and the reference summary template are kept unchanged.  This
script copies the template workbook and only fills the street data cells in the
generated copy.
"""

from __future__ import annotations

import math
import re
from copy import copy
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"

SOURCE_PREFIX = "1.2026-2027年早高峰时段各街道问题汇总台账"
TEMPLATE_SHEETS = ["Sheet1", "Sheet2"]
DEFAULT_OUTPUT_NAME = "（汇总）2026.5月各街道问题汇总.xlsx"

START_DATE = datetime(2026, 4, 20, 0, 0, 0)
END_DATE = datetime(2026, 5, 19, 23, 59, 59)


def clean_text(value) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", "").replace("\r", "")
    text = re.sub(r"\s+", "", text)
    return text.strip()


def metric_key(value) -> str:
    text = clean_text(value)
    text = text.replace("设桶", "摆桶")
    text = text.replace("未更新新国际", "未更新新国标")
    text = text.replace("未更新新國際", "未更新新国标")
    text = text.replace("容器标识不合格数", "容器标识不合格")
    text = text.replace("容器无标识或标识不合格数", "容器无标识或标识不合格")
    text = text.replace("无称重计量列表", "无称重计量")
    text = text.replace("无称重计量小程序", "无称重计量")
    text = text.replace("厨余、其他容器", "厨余和其他容器")
    text = text.replace("周边不洁", "周边不洁")
    return text


def street_key(value) -> str:
    text = clean_text(value)
    if not text:
        return ""
    aliases = {
        "西长安街": "西长安街街道",
        "什刹海街道": "什刹海街道",
        "广外": "广安门外街道",
        "广外街道": "广安门外街道",
        "广安门外": "广安门外街道",
        "广内": "广安门内街道",
        "广内街道": "广安门内街道",
        "广安门内": "广安门内街道",
    }
    if text in aliases:
        text = aliases[text]
    if not text.endswith("街道"):
        text = f"{text}街道"
    return text


def to_datetime(value) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return from_excel(value)
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                parsed = datetime.strptime(text, fmt)
                return datetime.combine(parsed.date(), time.min)
            except ValueError:
                pass
    return None


def parse_ui_date(value: str, end_of_day: bool = False) -> datetime:
    parsed = datetime.strptime(value, "%Y-%m-%d")
    if end_of_day:
        return datetime.combine(parsed.date(), time.max.replace(microsecond=0))
    return datetime.combine(parsed.date(), time.min)


def output_name_for(end_date: datetime) -> str:
    return f"（汇总）{end_date.year}.{end_date.month}月各街道问题汇总.xlsx"


def number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return 0.0


def find_source_file() -> Path:
    candidates = [
        path
        for path in DATA_DIR.glob("*.xlsx")
        if not path.name.startswith("~$") and path.name.startswith(SOURCE_PREFIX)
    ]
    if not candidates:
        raise FileNotFoundError(f"未找到源台账文件: {DATA_DIR / (SOURCE_PREFIX + '*.xlsx')}")
    return candidates[0]


def find_template_file() -> Path:
    for path in DATA_DIR.glob("*.xlsx"):
        if path.name.startswith("~$") or path.name.endswith("_生成.xlsx"):
            continue
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            if wb.sheetnames == TEMPLATE_SHEETS or find_summary_template_sheet(wb) is not None:
                return path
        finally:
            wb.close()
    raise FileNotFoundError("未找到包含 Sheet1/Sheet2 的汇总模板文件")


@dataclass
class SourceSheet:
    name: str
    headers: list[str]
    rows: list[tuple]
    indicator_indexes: list[int]
    date_index: int = 0
    street_index: int = 1
    rectified_index: int | None = None
    resident_vote_total_index: int | None = None
    resident_bin_total_index: int | None = None
    resident_station_group_index: int | None = None

    @property
    def header_map(self) -> dict[str, list[int]]:
        mapping: dict[str, list[int]] = defaultdict(list)
        for idx, header in enumerate(self.headers):
            if header:
                mapping[metric_key(header)].append(idx)
        return mapping

    def sum_metric(self, street: str, metric: str) -> float | None:
        indexes = self.header_map.get(metric_key(metric), [])
        if not indexes:
            return None
        total = 0.0
        wanted = street_key(street)
        for row in self.rows:
            if street_key(row[self.street_index]) != wanted:
                continue
            total += sum(number(row[idx]) for idx in indexes if idx < len(row))
        return total

    def sum_indexes(self, street: str, indexes: list[int]) -> float:
        total = 0.0
        wanted = street_key(street)
        for row in self.rows:
            if street_key(row[self.street_index]) != wanted:
                continue
            total += sum(number(row[idx]) for idx in indexes if idx < len(row))
        return total

    def issue_total(self, street: str) -> float:
        return self.sum_indexes(street, self.indicator_indexes)

    def row_count(self, street: str) -> int:
        wanted = street_key(street)
        return sum(1 for row in self.rows if street_key(row[self.street_index]) == wanted)

    def date_street_count(self, street: str) -> int:
        wanted = street_key(street)
        pairs = {
            to_datetime(row[self.date_index]).date()
            for row in self.rows
            if street_key(row[self.street_index]) == wanted and to_datetime(row[self.date_index]) is not None
        }
        return len(pairs)

    def default_resident_vote_total(self, street: str) -> int:
        return self.row_count(street) * 5

    def point_count(self, street: str, point_index: int = 2) -> int:
        wanted = street_key(street)
        points = {
            clean_text(row[point_index])
            for row in self.rows
            if len(row) > point_index and street_key(row[self.street_index]) == wanted and clean_text(row[point_index])
        }
        return len(points)

    def rectified_issue_totals(self, street: str) -> tuple[float, float]:
        if self.rectified_index is None:
            return 0.0, 0.0
        wanted = street_key(street)
        done = 0.0
        undone = 0.0
        for row in self.rows:
            if street_key(row[self.street_index]) != wanted:
                continue
            issue_count = sum(number(row[idx]) for idx in self.indicator_indexes if idx < len(row))
            rectified_value = number(row[self.rectified_index]) if self.rectified_index < len(row) else 0
            if rectified_value == 1:
                undone += issue_count
            else:
                done += issue_count
        return done, undone

    def rectified_record_totals(self, street: str) -> tuple[float, float]:
        if self.rectified_index is None:
            return 0.0, 0.0
        wanted = street_key(street)
        done = 0.0
        undone = 0.0
        for row in self.rows:
            if street_key(row[self.street_index]) != wanted:
                continue
            issue_count = sum(number(row[idx]) for idx in self.indicator_indexes if idx < len(row))
            if issue_count == 0:
                continue
            rectified_value = number(row[self.rectified_index]) if self.rectified_index < len(row) else 0
            if rectified_value == 1:
                undone += 1
            else:
                done += 1
        return done, undone


def load_source_sheet(ws, sheet_order: int, start_date: datetime, end_date: datetime) -> SourceSheet:
    header_row = 2 if sheet_order == 0 else 1
    headers = [clean_text(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    raw_rows = ws.iter_rows(min_row=header_row + 1, values_only=True)
    rows = []
    for row in raw_rows:
        checked_at = to_datetime(row[0] if row else None)
        if checked_at and start_date <= checked_at <= end_date:
            rows.append(row)

    # Excel column numbers converted to 0-based indexes.  These ranges come
    # from the ledger structure and exclude monthly-report helper columns.
    indicator_ranges_by_order = {
        0: [(6, 31), (33, 42), (44, 45), (47, 58)],
        1: [(4, 36)],
        2: [(4, 31)],
        3: [(4, 33)],
        4: [(5, 23)],
        5: [(4, 13)],
        6: [(4, 8)],
    }
    ranges = indicator_ranges_by_order[sheet_order]
    indicator_indexes = [col - 1 for start, end in ranges for col in range(start, end + 1)]

    def first_header_contains(keyword: str) -> int | None:
        return next((idx for idx, header in enumerate(headers) if keyword in header), None)

    sheet = SourceSheet(
        name=ws.title,
        headers=headers,
        rows=rows,
        indicator_indexes=indicator_indexes,
        rectified_index=first_header_contains("整改情况"),
    )
    if sheet_order == 0:
        sheet.resident_station_group_index = 31
        sheet.resident_bin_total_index = 42
        sheet.resident_vote_total_index = 45
    return sheet


def load_source_data(source_file: Path, start_date: datetime = START_DATE, end_date: datetime = END_DATE) -> dict[str, SourceSheet]:
    wb = load_workbook(source_file, data_only=True, read_only=True)
    data: dict[str, SourceSheet] = {}
    for order, name in enumerate(wb.sheetnames):
        data[name] = load_source_sheet(wb[name], order, start_date, end_date)
    return data


def first_sheet(data: dict[str, SourceSheet], keyword: str) -> SourceSheet:
    return next(sheet for name, sheet in data.items() if keyword in name)


def rows_with_category(ws) -> dict[int, str]:
    current = ""
    categories = {}
    for row in range(2, ws.max_row + 1):
        value = clean_text(ws.cell(row, 2).value)
        if value:
            current = value
        categories[row] = current
    return categories


def same_category_detail_rows(ws, categories: dict[int, str], row: int) -> list[int]:
    category = categories[row]
    detail_rows = []
    for candidate in range(row + 1, ws.max_row + 1):
        if categories[candidate] != category:
            break
        label = clean_text(ws.cell(candidate, 3).value)
        if label and label != "合计":
            detail_rows.append(candidate)
    return detail_rows


def build_value_resolver(data: dict[str, SourceSheet], ws, categories: dict[int, str]) -> Callable:
    resident = first_sheet(data, "居民")
    social = first_sheet(data, "社会")
    restaurant = first_sheet(data, "餐饮")
    transfer = first_sheet(data, "中转")

    resident_aliases = {
        "无桶站公示牌或桶站公示牌设置不合格": "无桶站公示牌或桶站公示牌不合格",
        "大件、装修垃圾投放点周边环境脏乱": "桶站周边不洁",
        "小区内无大件垃圾托底上门回收渠道公示或告知": "未公示收集点位和收集时间",
        "无灭蝇措施": "无灭蚊蝇设施",
        "车辆未密闭、破损滴漏脏污，标志与车牌不清晰": "车辆未密闭、破损滴漏脏污，无标识、标志与车牌不清晰",
        "散桶": "散桶",
        "站外摆桶": "站外摆桶",
    }
    social_aliases = {
        "容器破损、脏污等": "容器容器破损",
        "食品加工区厨余、其他容器设置不齐": "食品加工区厨余和其他容器设置不齐",
        "无称重计量列表": "无称重计量",
        "容器标识不合格": "容器标识不合格",
    }
    restaurant_aliases = {
        "桶站周边不洁": "桶站周边不洁",
        "容器无标识或标识不合格": "容器无标识或标识不合格",
        "食品加工区厨余、其他容器设置不齐": "食品加工区厨余和其他容器设置不齐",
        "无称重计量小程序": "无称重计量",
    }
    transfer_aliases = {
        "无消防安全水源": "无消防安全水源",
        "周边环境脏乱": "周边环境脏乱",
    }

    def source_for(row: int) -> SourceSheet | None:
        category = categories[row]
        if category in ("分类设施建设达标情况", "分类设施管理达标情况", "居民自主投放情况", "纯净率", "值守率", "检查小区桶站组数", "居民小区问题整改情况"):
            return resident
        if category == "社会单位检查情况":
            return social
        if category == "餐饮单位检查情况":
            return restaurant
        if category == "中转站":
            return transfer
        return None

    def mapped_metric(sheet: SourceSheet, metric: str) -> str:
        key = metric_key(metric)
        if sheet is resident:
            return resident_aliases.get(key, key)
        if sheet is social:
            return social_aliases.get(key, key)
        if sheet is restaurant:
            return restaurant_aliases.get(key, key)
        if sheet is transfer:
            return transfer_aliases.get(key, key)
        return key

    source_metric_owner: dict[tuple[str, int, str], tuple[int, bool]] = {}
    for owner_row in range(2, ws.max_row + 1):
        owner_metric = clean_text(ws.cell(owner_row, 3).value)
        if not owner_metric or owner_metric == "合计":
            continue
        owner_category = categories[owner_row]
        owner_sheet = source_for(owner_row)
        if owner_sheet is None:
            continue
        owner_source_metric = mapped_metric(owner_sheet, owner_metric)
        owner_is_exact = metric_key(owner_metric) == owner_source_metric
        owner_key = (owner_category, id(owner_sheet), owner_source_metric)
        current = source_metric_owner.get(owner_key)
        if current is None or (owner_is_exact and not current[1]):
            source_metric_owner[owner_key] = (owner_row, owner_is_exact)

    def value(row: int, street: str, cache: dict[tuple[int, str], float | None], missing: list[str]) -> float | None:
        cache_key = (row, street_key(street))
        if cache_key in cache:
            return cache[cache_key]

        metric = clean_text(ws.cell(row, 3).value)
        category = categories[row]
        sheet = source_for(row)
        result: float | None

        if metric == "合计":
            result = sum(value(detail_row, street, cache, missing) or 0 for detail_row in same_category_detail_rows(ws, categories, row))
        elif category == "居民小区问题整改情况" and metric == "居民小区、胡同整改率":
            result = 1
        elif category == "居民小区问题整改情况" and metric == "检查数":
            result = resident.date_street_count(street)
        elif category == "居民小区问题整改情况" and metric == "已整改数":
            result = resident.date_street_count(street)
        elif category == "居民小区问题整改情况" and metric == "未整改数":
            result = 0
        elif category == "居民自主投放情况" and metric == "投放准确率":
            bad = resident.sum_metric(street, "居民自主投放不准确") or 0
            total = resident.default_resident_vote_total(street)
            result = 1 - bad / total if total else 1
        elif category == "纯净率":
            bad = resident.sum_metric(street, "桶内分类不纯净") or 0
            total = resident.sum_indexes(street, [resident.resident_bin_total_index]) if resident.resident_bin_total_index is not None else 0
            result = 1 - bad / total if total else 1
        elif category == "值守率":
            bad = resident.sum_metric(street, "无人值守") or 0
            total = resident.sum_indexes(street, [resident.resident_station_group_index]) if resident.resident_station_group_index is not None else 0
            result = 1 - bad / total if total else 1
        elif category == "检查小区桶站组数":
            result = resident.sum_indexes(street, [resident.resident_station_group_index]) if resident.resident_station_group_index is not None else None
        elif category == "社会单位检查情况" and metric == "社会单位检查问题数":
            result = social.issue_total(street)
        elif category == "餐饮单位检查情况" and metric == "餐饮单位检查问题数":
            result = restaurant.issue_total(street)
        elif category == "市级检查情况" and metric == "检查小区数":
            result = resident.point_count(street)
        elif category == "市级检查情况" and metric == "检查社会单位数":
            result = social.point_count(street)
        elif category == "市级检查情况" and metric == "市级检查问题数":
            result = resident.issue_total(street) + social.issue_total(street) + restaurant.issue_total(street)
        elif sheet is not None:
            source_metric = mapped_metric(sheet, metric)
            owner = source_metric_owner.get((category, id(sheet), source_metric))
            if owner is not None and owner[0] != row:
                result = 0
            else:
                result = sheet.sum_metric(street, source_metric)
            if result is None:
                missing.append(f"第{row}行 [{category}] {metric} -> {sheet.name}")
                result = 0
        else:
            result = 0
            missing.append(f"第{row}行 [{category}] {metric} -> 未确定源表")

        cache[cache_key] = result
        return result

    return value


def find_summary_template_sheet(wb):
    if "Sheet1" in wb.sheetnames:
        return wb["Sheet1"]
    for ws in wb.worksheets:
        street_headers = [
            clean_text(ws.cell(1, col).value)
            for col in range(4, ws.max_column + 1)
            if street_key(ws.cell(1, col).value)
        ]
        category_hits = 0
        metric_hits = 0
        for row in range(2, min(ws.max_row, 120) + 1):
            category = clean_text(ws.cell(row, 2).value)
            metric = clean_text(ws.cell(row, 3).value)
            if category in {
                "分类设施建设达标情况", "分类设施管理达标情况", "中转站",
                "社会单位检查情况", "餐饮单位检查情况",
            }:
                category_hits += 1
            if metric:
                metric_hits += 1
        if len(street_headers) >= 5 and category_hits >= 2 and metric_hits >= 10:
            return ws
    return wb[wb.sheetnames[0]] if wb.sheetnames else None


def fill_sheet1(wb, data: dict[str, SourceSheet]) -> list[str]:
    ws = find_summary_template_sheet(wb)
    if ws is None:
        raise ValueError("汇总模板中未找到可写入的工作表")
    street_columns = [
        (col, clean_text(ws.cell(1, col).value))
        for col in range(4, ws.max_column + 1)
        if clean_text(ws.cell(1, col).value)
    ]
    categories = rows_with_category(ws)
    resolve_value = build_value_resolver(data, ws, categories)
    missing: list[str] = []
    cache: dict[tuple[int, str], float | None] = {}

    for row in range(2, ws.max_row + 1):
        category = categories[row]
        if category == "市级检查情况":
            continue
        metric = clean_text(ws.cell(row, 3).value)
        if not metric and category != "检查小区桶站组数":
            continue
        for col, street in street_columns:
            value = resolve_value(row, street, cache, missing)
            cell = ws.cell(row, col)
            if value is None:
                cell.value = 0
                continue
            cell.value = round(value, 4) if isinstance(value, float) and not value.is_integer() else int(value)

    return sorted(set(missing))


def category_bounds(ws, category: str) -> tuple[int, int] | None:
    wanted = clean_text(category)
    current = ""
    start = None
    end = None
    for row in range(2, ws.max_row + 1):
        category_value = clean_text(ws.cell(row, 2).value)
        if category_value:
            current = category_value
        if current == wanted:
            if start is None:
                start = row
            end = row
        elif start is not None:
            break
    if start is None or end is None:
        return None
    return start, end


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def append_missing_issue_rows(wb, data: dict[str, SourceSheet]) -> list[str]:
    ws = find_summary_template_sheet(wb)
    if ws is None:
        raise ValueError("汇总模板中未找到可写入的工作表")
    street_columns = [
        (col, clean_text(ws.cell(1, col).value))
        for col in range(4, ws.max_column + 1)
        if clean_text(ws.cell(1, col).value)
    ]
    issue_sections = [
        ("社会单位检查情况", first_sheet(data, "社会"), {
            "容器容器破损": "容器破损、脏污等",
            "食品加工区厨余和其他容器设置不齐": "食品加工区厨余、其他容器设置不齐",
            "无称重计量": "无称重计量列表",
            "容器标识不合格": "容器标识不合格数",
        }),
        ("餐饮单位检查情况", first_sheet(data, "餐饮"), {
            "容器无标识或标识不合格": "容器无标识或标识不合格数",
            "食品加工区厨余和其他容器设置不齐": "食品加工区厨余、其他容器设置不齐",
            "无称重计量": "无称重计量小程序",
        }),
    ]
    inserted: list[str] = []
    for category, sheet, aliases in issue_sections:
        bounds = category_bounds(ws, category)
        if bounds is None:
            continue
        start, end = bounds
        existing = {
            metric_key(ws.cell(row, 3).value)
            for row in range(start, end + 1)
            if clean_text(ws.cell(row, 3).value)
        }
        reusable_rows = [
            row
            for row in range(start, end + 1)
            if clean_text(ws.cell(row, 3).value)
            and clean_text(ws.cell(row, 3).value) not in {"社会单位检查问题数", "餐饮单位检查问题数"}
            and sum(number(ws.cell(row, col).value) for col, _ in street_columns) == 0
        ]
        source_metrics = []
        seen = set()
        for idx in sheet.indicator_indexes:
            if idx >= len(sheet.headers):
                continue
            source_metric = clean_text(sheet.headers[idx])
            if not source_metric or source_metric in {"合计", "整改情况（已整改：0，未整改：1）"}:
                continue
            key = metric_key(source_metric)
            if key in seen:
                continue
            seen.add(key)
            display_metric = aliases.get(key, source_metric)
            if metric_key(display_metric) in existing:
                continue
            values = {street: sheet.sum_metric(street, source_metric) or 0 for _, street in street_columns}
            if sum(values.values()) <= 0:
                continue
            source_metrics.append((display_metric, values))

        for (display_metric, values), target_row in zip(source_metrics, reusable_rows):
            ws.cell(target_row, 3).value = display_metric
            for col, street in street_columns:
                value = values[street]
                ws.cell(target_row, col).value = round(value, 4) if isinstance(value, float) and not value.is_integer() else int(value)
            inserted.append(f"{category} / {display_metric}")
            existing.add(metric_key(display_metric))
        if len(source_metrics) > len(reusable_rows):
            for display_metric, _ in source_metrics[len(reusable_rows):]:
                inserted.append(f"{category} / {display_metric}（未写入：模板无可复用空行）")
    return inserted


def generate_summary(
    source_file: Path,
    template_file: Path,
    start_date: datetime,
    end_date: datetime,
    output_file: Path | None = None,
) -> dict:
    output_file = output_file or DATA_DIR / output_name_for(end_date)
    source_data = load_source_data(source_file, start_date, end_date)
    wb = load_workbook(template_file, data_only=False)
    missing = fill_sheet1(wb, source_data)
    inserted_rows = append_missing_issue_rows(wb, source_data)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    return {
        "output_file": output_file,
        "missing": missing,
        "inserted_rows": inserted_rows,
        "row_counts": {name: len(sheet.rows) for name, sheet in source_data.items()},
        "start_date": start_date,
        "end_date": end_date,
    }


def main() -> None:
    source_file = find_source_file()
    template_file = find_template_file()
    output_file = DATA_DIR / DEFAULT_OUTPUT_NAME

    print(f"源台账: {source_file}")
    print(f"模板: {template_file}")
    print(f"输出: {output_file}")
    print(f"日期范围: {START_DATE:%Y-%m-%d} 至 {END_DATE:%Y-%m-%d}")

    result = generate_summary(source_file, template_file, START_DATE, END_DATE, output_file)

    print("生成完成。")
    for name, count in result["row_counts"].items():
        print(f"- {name}: 筛选后 {count} 行")
    if result["missing"]:
        print("\n以下模板行没有找到明确源列，生成文件中对应数据格已留空，请补充映射后再生成：")
        for item in result["missing"]:
            print(f"- {item}")
    else:
        print("所有可填充模板行均已映射。")


if __name__ == "__main__":
    main()
