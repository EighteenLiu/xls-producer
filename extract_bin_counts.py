# -*- coding: utf-8 -*-
"""Extract bin/container counts from the raw issue export and update rates.

The source file is an old ``.xls`` export whose useful headers span rows 1-2:
street is column D and the raw issue text is column F ("具体问题").
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import win32com.client as win32
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SOURCE = BASE_DIR / "data" / "0420-0519.xls"
DEFAULT_CONVERTED = BASE_DIR / "web_work" / "outputs" / "0420-0519_converted.xlsx"
DEFAULT_AUDIT = BASE_DIR / "web_work" / "outputs" / "垃圾桶数量提取校对表.xlsx"

STREETS = [
    "德胜街道",
    "什刹海街道",
    "西长安街街道",
    "大栅栏街道",
    "天桥街道",
    "新街口街道",
    "金融街街道",
    "椿树街道",
    "陶然亭街道",
    "展览路街道",
    "月坛街道",
    "广内街道",
    "牛街街道",
    "白纸坊街道",
    "广外街道",
]

CN_NUMBERS = {
    "零": 0,
    "一": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
    "十": 10,
}


@dataclass
class CountRecord:
    row: int
    record_id: str
    street: str
    point_type: str
    point_name: str
    issue_text: str
    count_type: str
    count: int
    matched_text: str
    included: str
    indicator_2: str
    indicator_3: str


def chinese_to_int(value: str) -> int | None:
    value = value.strip()
    if not value:
        return None
    if value.isdigit():
        return int(value)
    if value in CN_NUMBERS:
        return CN_NUMBERS[value]
    if "十" in value:
        left, _, right = value.partition("十")
        tens = CN_NUMBERS.get(left, 1) if left else 1
        ones = CN_NUMBERS.get(right, 0) if right else 0
        return tens * 10 + ones
    return None


def convert_xls_to_xlsx(source: Path, output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    if source.suffix.lower() == ".xlsx":
        return source
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(source.resolve()))
        wb.SaveAs(str(output.resolve()), FileFormat=51)
        wb.Close(False)
    finally:
        excel.Quit()
    return output


def context(text: str, start: int, end: int, pad: int = 18) -> str:
    return text[max(0, start - pad) : min(len(text), end + pad)]


def extract_counts(text: str) -> list[tuple[str, int, str, str]]:
    records: list[tuple[str, int, str, str]] = []
    seen: set[tuple[str, int, str]] = set()

    def add(count_type: str, count: int | None, matched: str, included: str) -> None:
        if count is None:
            return
        key = (count_type, count, matched)
        if key in seen:
            return
        seen.add(key)
        records.append((count_type, count, matched, included))

    # Strong denominator signals: total checked containers/bins.
    total_container_patterns = [
        r"(?:共查|共检查|一共检查了|今天一共检查了|检查了)\s*([0-9一二两三四五六七八九十]+)\s*个\s*(?:桶|容器)",
        r"[（(]\s*([0-9一二两三四五六七八九十]+)\s*个\s*桶\s*[）)]",
        r"共\s*([0-9一二两三四五六七八九十]+)\s*个\s*(?:桶|容器)",
    ]
    for pattern in total_container_patterns:
        for match in re.finditer(pattern, text):
            add("容器总数", chinese_to_int(match.group(1)), match.group(0), "是")

    # Strong denominator signals: station groups.
    group_patterns = [
        r"(?:只剩下?|只有|剩余|共|检查了|一共检查了)?\s*([0-9一二两三四五六七八九十]+)\s*组\s*(?:桶站|垃圾桶站|垃圾桶)",
    ]
    for pattern in group_patterns:
        for match in re.finditer(pattern, text):
            add("桶站组数", chinese_to_int(match.group(1)), match.group(0), "是")

    # Audit-only signals: specific bin counts mentioned in problem text. These
    # are useful for manual review, but are not safe denominators by default.
    bin_count_patterns = [
        r"([0-9一二两三四五六七八九十]+)\s*个\s*(?:厨余|其他|可回收物|有害)?垃圾桶",
        r"([0-9一二两三四五六七八九十]+)\s*个\s*(?:厨余|其他|可回收物|有害)?桶",
        r"(?:厨余|其他|可回收物|有害)?垃圾桶\s*([0-9一二两三四五六七八九十]+)\s*个",
    ]
    for pattern in bin_count_patterns:
        for match in re.finditer(pattern, text):
            matched = context(text, match.start(), match.end())
            if any(existing[2] in matched or matched in existing[2] for existing in records):
                continue
            add("垃圾桶个数候选", chinese_to_int(match.group(1)), matched, "否")

    # Parentheses containing bin-related words are included for review even
    # when the exact number is not confidently parsed.
    for match in re.finditer(r"[（(][^）)]*(?:桶|容器|桶站)[^）)]*[）)]", text):
        matched = match.group(0)
        if any(existing[2] == matched for existing in records):
            continue
        num_match = re.search(r"([0-9一二两三四五六七八九十]+)", matched)
        add("括号内桶数量候选", chinese_to_int(num_match.group(1)) if num_match else 0, matched, "否")

    return records


def extract_records(source_xlsx: Path) -> list[CountRecord]:
    wb = load_workbook(source_xlsx, data_only=True)
    ws = wb.active
    records: list[CountRecord] = []
    for row in range(3, ws.max_row + 1):
        issue_text = str(ws.cell(row, 6).value or "").strip()
        if not issue_text:
            continue
        street = str(ws.cell(row, 4).value or "").strip()
        if street not in STREETS:
            continue
        for count_type, count, matched, included in extract_counts(issue_text):
            records.append(
                CountRecord(
                    row=row,
                    record_id=str(ws.cell(row, 1).value or ""),
                    street=street,
                    point_type=str(ws.cell(row, 3).value or ""),
                    point_name=str(ws.cell(row, 5).value or ""),
                    issue_text=issue_text,
                    count_type=count_type,
                    count=count,
                    matched_text=matched,
                    included=included,
                    indicator_2=str(ws.cell(row, 60).value or ""),
                    indicator_3=str(ws.cell(row, 61).value or ""),
                )
            )
    return records


def totals_by_street(records: list[CountRecord]) -> dict[str, dict[str, int]]:
    totals = {street: {"容器总数": 0, "桶站组数": 0} for street in STREETS}
    for record in records:
        if record.included != "是":
            continue
        if record.count_type in totals[record.street]:
            totals[record.street][record.count_type] += record.count
    return totals


def write_audit_workbook(records: list[CountRecord], output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "明细校对"
    headers = [
        "原始行号",
        "编号",
        "街道",
        "点位类型",
        "点位名称",
        "数量类型",
        "提取数量",
        "计入分母",
        "匹配片段",
        "具体问题原文",
        "2级指标",
        "3级指标",
    ]
    ws.append(headers)
    for record in records:
        ws.append(
            [
                record.row,
                record.record_id,
                record.street,
                record.point_type,
                record.point_name,
                record.count_type,
                record.count,
                record.included,
                record.matched_text,
                record.issue_text,
                record.indicator_2,
                record.indicator_3,
            ]
        )

    summary = wb.create_sheet("街道汇总")
    summary.append(["街道", "容器总数", "桶站组数"])
    totals = totals_by_street(records)
    for street in STREETS:
        summary.append([street, totals[street]["容器总数"], totals[street]["桶站组数"]])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output)
    return output


def update_summary_rates(summary_file: Path, records: list[CountRecord], output: Path) -> Path:
    wb = load_workbook(summary_file)
    ws = wb["Sheet1"]
    totals = totals_by_street(records)
    street_columns = {
        str(ws.cell(1, col).value or "").strip(): col
        for col in range(4, ws.max_column + 1)
        if str(ws.cell(1, col).value or "").strip()
    }
    for street, col in street_columns.items():
        container_total = totals.get(street, {}).get("容器总数", 0)
        station_total = totals.get(street, {}).get("桶站组数", 0)
        impure = ws.cell(41, col).value or 0
        unattended = ws.cell(33, col).value or 0
        ws.cell(53, col).value = round(1 - impure / container_total, 4) if container_total else (1 if not impure else 0)
        ws.cell(54, col).value = round(1 - unattended / station_total, 4) if station_total else (1 if not unattended else 0)
        ws.cell(55, col).value = station_total
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract garbage-bin counts and update summary rates.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--converted", type=Path, default=DEFAULT_CONVERTED)
    parser.add_argument("--audit-output", type=Path, default=DEFAULT_AUDIT)
    parser.add_argument("--summary-input", type=Path)
    parser.add_argument("--summary-output", type=Path)
    args = parser.parse_args()

    source_xlsx = convert_xls_to_xlsx(args.source, args.converted)
    records = extract_records(source_xlsx)
    audit_path = write_audit_workbook(records, args.audit_output)
    print(f"校对表: {audit_path}")

    if args.summary_input:
        summary_output = args.summary_output or args.summary_input.with_name(args.summary_input.stem + "_更新纯净率值守率.xlsx")
        updated_path = update_summary_rates(args.summary_input, records, summary_output)
        print(f"更新后的汇总表: {updated_path}")


if __name__ == "__main__":
    main()
