# -*- coding: utf-8 -*-
"""Local desktop UI for generating the monthly street summary workbook.

Run:
    python app.py

For the old browser UI:
    python app.py --web
"""

from __future__ import annotations

import warnings

warnings.filterwarnings("ignore", category=DeprecationWarning, message="'cgi' is deprecated.*")

import cgi
import json
import mimetypes
import re
import shutil
import sys
import threading
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import Workbook, load_workbook

from generate_monthly_summary import clean_text, generate_summary, metric_key, number, output_name_for, parse_ui_date, street_key, to_datetime
from report_generator import generate_reports


BASE_DIR = Path(__file__).resolve().parent
WORK_DIR = BASE_DIR / "web_work"
UPLOAD_DIR = WORK_DIR / "uploads"
DEFAULT_OUTPUT_DIR = WORK_DIR / "outputs"
PROCESS_DIR = WORK_DIR / "process"
HOST = "127.0.0.1"
PORT = 8000

DOWNLOAD_FILES: dict[str, Path] = {}


INDEX_HTML = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>街道问题汇总表生成</title>
  <style>
    :root {
      --bg: #f6f7f9;
      --panel: #ffffff;
      --text: #20242a;
      --muted: #667085;
      --line: #d9dee7;
      --accent: #1f7a6d;
      --accent-dark: #176258;
      --warn: #9a5b12;
      --error: #b42318;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Microsoft YaHei", "Segoe UI", Arial, sans-serif;
      background: var(--bg);
      color: var(--text);
    }
    body.modal-open {
      overflow: hidden;
    }
    main {
      width: min(1040px, calc(100vw - 32px));
      margin: 28px auto 40px;
    }
    header {
      display: flex;
      align-items: end;
      justify-content: space-between;
      gap: 20px;
      margin-bottom: 18px;
    }
    h1 {
      margin: 0;
      font-size: 26px;
      letter-spacing: 0;
      font-weight: 700;
    }
    .subtitle {
      margin: 8px 0 0;
      color: var(--muted);
      font-size: 14px;
    }
    .home-view {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 22px;
    }
    .entry-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 16px;
    }
    .entry-card {
      display: block;
      width: 100%;
      min-height: 150px;
      padding: 22px;
      text-align: left;
      background: #fff;
      color: var(--text);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 1px 2px rgba(16, 24, 40, .04);
    }
    .entry-card:hover {
      background: #f8fbfa;
      border-color: #9fc8c0;
    }
    .entry-title {
      display: block;
      font-size: 20px;
      font-weight: 700;
      margin-bottom: 10px;
    }
    .entry-desc {
      display: block;
      color: var(--muted);
      font-size: 14px;
      font-weight: 400;
      line-height: 1.6;
      white-space: normal;
    }
    .view[hidden], .home-view[hidden] {
      display: none;
    }
    .view {
      position: fixed;
      inset: 0;
      z-index: 20;
      overflow-y: auto;
      padding: 28px 16px;
      background: rgba(16, 24, 40, .45);
    }
    .modal-panel {
      width: min(1120px, calc(100vw - 32px));
      margin: 0 auto;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 18px 48px rgba(16, 24, 40, .22);
      padding: 20px;
    }
    .view-header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 14px;
      margin-bottom: 14px;
    }
    .view-header h2 {
      margin: 0;
      font-size: 20px;
    }
    .layout {
      display: grid;
      grid-template-columns: 1.05fr .95fr;
      gap: 18px;
    }
    section, aside {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 20px;
    }
    .grid {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
    }
    label {
      display: block;
      font-size: 13px;
      color: #344054;
      margin-bottom: 8px;
      font-weight: 600;
    }
    input[type="file"], input[type="date"], input[type="text"], select {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      color: var(--text);
      padding: 11px 12px;
      font-size: 14px;
    }
    input[type="file"] { min-height: 44px; }
    .field { margin-bottom: 16px; }
    .path-row {
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 10px;
      align-items: center;
    }
    .hint {
      margin-top: 6px;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.5;
    }
    .actions {
      display: flex;
      align-items: center;
      gap: 12px;
      margin-top: 4px;
    }
    button, .download {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
      min-height: 42px;
      border: 0;
      border-radius: 6px;
      padding: 0 18px;
      background: var(--accent);
      color: #fff;
      font-size: 14px;
      font-weight: 700;
      cursor: pointer;
      text-decoration: none;
      white-space: nowrap;
    }
    button:hover, .download:hover { background: var(--accent-dark); }
    .secondary {
      min-height: 42px;
      background: #eef4f3;
      color: var(--accent-dark);
      border: 1px solid #b7d4ce;
    }
    .secondary:hover { background: #deebe8; }
    button:disabled {
      opacity: .6;
      cursor: not-allowed;
    }
    .output-name {
      padding: 12px;
      border-radius: 6px;
      border: 1px dashed #b8c4d6;
      background: #fbfcfe;
      font-family: Consolas, "Microsoft YaHei", monospace;
      font-size: 13px;
      overflow-wrap: anywhere;
    }
    .status {
      min-height: 42px;
      margin-top: 16px;
      padding: 12px;
      border-radius: 6px;
      background: #f2f4f7;
      color: var(--muted);
      font-size: 14px;
      line-height: 1.6;
      white-space: pre-wrap;
    }
    .status.ok { color: #075e54; background: #eaf6f2; }
    .status.warn { color: var(--warn); background: #fff7e8; }
    .status.error { color: var(--error); background: #fff0ee; }
    .summary-title {
      margin: 0 0 12px;
      font-size: 16px;
    }
    .kv {
      display: grid;
      grid-template-columns: 112px 1fr;
      gap: 8px 12px;
      font-size: 14px;
      margin-bottom: 18px;
    }
    .kv span:nth-child(odd) { color: var(--muted); }
    .counts {
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 8px 12px;
      font-size: 14px;
      padding-top: 12px;
      border-top: 1px solid var(--line);
    }
    .missing {
      margin: 12px 0 0;
      padding-left: 18px;
      color: var(--warn);
      font-size: 13px;
      line-height: 1.6;
    }
    .report-block {
      grid-column: 1;
      margin-top: 18px;
    }
    .layout > aside {
      grid-column: 2;
      grid-row: 1 / span 3;
    }
    .download-list {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-top: 12px;
    }
    .notes {
      margin-top: 18px;
      padding-top: 14px;
      border-top: 1px solid var(--line);
      color: var(--muted);
      font-size: 13px;
      line-height: 1.7;
    }
    .notes p {
      margin: 0 0 8px;
    }
    @media (max-width: 760px) {
      header, .layout, .grid, .entry-grid, .view-header { display: block; }
      .view { padding: 14px 10px; }
      .modal-panel {
        width: 100%;
        padding: 14px;
      }
      .layout > aside { display: block; }
      aside { margin-top: 18px; }
      .entry-card { margin-bottom: 12px; }
      .view-header .secondary { margin-top: 12px; }
      .path-row { grid-template-columns: 1fr; }
      .actions { align-items: stretch; flex-direction: column; }
      button, .download { width: 100%; }
    }
  </style>
</head>
<body>
  <main>
    <header>
      <div>
        <h1>月报工具</h1>
        <p class="subtitle">选择要处理的任务，再上传对应模板和台账文件。</p>
      </div>
    </header>
    <div id="homeView" class="home-view">
      <div class="entry-grid">
        <button id="openSummary" class="entry-card" type="button">
          <span class="entry-title">生成汇总表</span>
          <span class="entry-desc">进入后可生成月报汇总表，也可以续写容器总数、纯净率和市级检查情况。</span>
        </button>
        <button id="openReport" class="entry-card" type="button">
          <span class="entry-title">生成工作报告</span>
          <span class="entry-desc">上传汇总表、源台账和两个 Word 模板，生成两个工作报告文档。</span>
        </button>
      </div>
    </div>
    <div id="summaryView" class="view" hidden>
      <div class="modal-panel" role="dialog" aria-modal="true">
      <div class="view-header">
        <h2>生成汇总表</h2>
        <button class="secondary back-home" type="button">返回入口</button>
      </div>
      <div class="layout">
      <section>
        <h2 class="summary-title">生成月报汇总表</h2>
        <form id="form">
          <div class="field">
            <label for="template">汇总 xlsx 模板</label>
            <input id="template" name="template" type="file" accept=".xlsx" required />
          </div>
          <div class="field">
            <label for="source">数据源台账</label>
            <input id="source" name="source" type="file" accept=".xlsx" required />
          </div>
          <div class="grid">
            <div class="field">
              <label for="start_date">开始日期</label>
              <input id="start_date" name="start_date" type="date" value="2026-04-20" required />
            </div>
            <div class="field">
              <label for="end_date">结束日期</label>
              <input id="end_date" name="end_date" type="date" value="2026-05-19" required />
            </div>
          </div>
          <div class="field">
            <label for="output_dir">文件保存地址</label>
            <div class="path-row">
              <input id="output_dir" name="output_dir" type="text" value="D:/pycharm/xls-producer/web_work/outputs" required />
              <button id="chooseOutputDir" class="secondary" type="button">选择文件夹</button>
            </div>
            <div class="hint">生成文件会直接保存到该文件夹，也可以手动粘贴路径。</div>
          </div>
          <div class="field">
            <label>生成文件名</label>
            <div class="output-name" id="outputName">（汇总）2026.5月各街道问题汇总.xlsx</div>
          </div>
          <div class="actions">
            <button id="submit" type="submit">生成汇总表</button>
            <a id="download" class="download" href="#" hidden>下载结果</a>
          </div>
          <div id="status" class="status">等待上传文件。</div>
        </form>
      </section>
      <section class="report-block">
        <h2 class="summary-title">更新容器总数与纯净率</h2>
        <form id="purityForm">
          <div class="field">
            <label for="purity_source">上传区：带容器总数的原始台账</label>
            <input id="purity_source" name="source" type="file" accept=".xlsx,.xls" required />
            <div class="hint">支持 xlsx / xls。后台会识别“有1个垃圾桶、1组桶、1个桶”等具体问题列，并在 web_work/process 中生成对照表。</div>
          </div>
          <div class="field">
            <label for="purity_summary">待修改文件区：已生成的月报汇总表</label>
            <input id="purity_summary" name="summary" type="file" accept=".xlsx" required />
            <div class="hint">只更新该表中的容器总数与桶内分类纯净率，其他内容保持不变。</div>
          </div>
          <div class="grid">
            <div class="field">
              <label for="purity_start_date">开始日期</label>
              <input id="purity_start_date" name="start_date" type="date" value="2026-04-20" required />
            </div>
            <div class="field">
              <label for="purity_end_date">结束日期</label>
              <input id="purity_end_date" name="end_date" type="date" value="2026-05-19" required />
            </div>
          </div>
          <div class="field">
            <label for="purity_output_dir">输出区：更新后汇总表保存地址</label>
            <div class="path-row">
              <input id="purity_output_dir" name="output_dir" type="text" value="D:/pycharm/xls-producer/web_work/outputs" required />
              <button id="choosePurityOutputDir" class="secondary" type="button">选择文件夹</button>
            </div>
            <div class="hint">默认使用上传汇总表的原文件名保存到该目录；同名文件会被覆盖。</div>
          </div>
          <div class="actions">
            <button id="puritySubmit" type="submit">更新容器总数与纯净率</button>
          </div>
          <div id="purityStatus" class="status">等待上传原始台账和待修改汇总表。</div>
          <div id="purityDownloads" class="download-list"></div>
        </form>
      </section>
      <section class="report-block">
        <h2 class="summary-title">更新市级检查情况</h2>
        <form id="cityForm">
          <div class="field">
            <label for="city_summary">待修改文件区：已生成的月报汇总表</label>
            <input id="city_summary" name="summary" type="file" accept=".xlsx" required />
            <div class="hint">只更新“市级检查情况”三行，其他内容保持不变。</div>
          </div>
          <div class="grid">
            <div class="field">
              <label for="city_resident">市级小区汇总表</label>
              <input id="city_resident" name="resident_summary" type="file" accept=".xlsx" required />
            </div>
            <div class="field">
              <label for="city_social">市级社会汇总表</label>
              <input id="city_social" name="social_summary" type="file" accept=".xlsx" required />
            </div>
          </div>
          <div class="grid">
            <div class="field">
              <label for="city_start_date">开始日期</label>
              <input id="city_start_date" name="start_date" type="date" value="2026-04-20" required />
            </div>
            <div class="field">
              <label for="city_end_date">结束日期</label>
              <input id="city_end_date" name="end_date" type="date" value="2026-05-19" required />
            </div>
          </div>
          <div class="field">
            <label for="city_output_dir">输出区：更新后汇总表保存地址</label>
            <div class="path-row">
              <input id="city_output_dir" name="output_dir" type="text" value="D:/pycharm/xls-producer/web_work/outputs" required />
              <button id="chooseCityOutputDir" class="secondary" type="button">选择文件夹</button>
            </div>
            <div class="hint">默认使用上传汇总表的原文件名保存到该目录；同名文件会被覆盖。</div>
          </div>
          <div class="actions">
            <button id="citySubmit" type="submit">更新市级检查情况</button>
          </div>
          <div id="cityStatus" class="status">等待上传月汇总表和两个市级汇总表。</div>
          <div id="cityDownloads" class="download-list"></div>
        </form>
      </section>
      <section class="report-block">
        <h2 class="summary-title">更新中转站情况</h2>
        <form id="transferForm">
          <div class="field">
            <label for="transfer_ledger">上传区：中转站台账</label>
            <input id="transfer_ledger" name="transfer_ledger" type="file" accept=".xlsx,.xls" required />
            <div class="hint">支持 xlsx / xls。程序会按所选时间段筛选日期，并按街道汇总中转站问题列。</div>
          </div>
          <div class="field">
            <label for="transfer_summary">待修改文件区：已生成的月报汇总表</label>
            <input id="transfer_summary" name="summary" type="file" accept=".xlsx" required />
            <div class="hint">只更新汇总表中的“中转站”相关行，其他内容保持不变。</div>
          </div>
          <div class="grid">
            <div class="field">
              <label for="transfer_start_date">开始日期</label>
              <input id="transfer_start_date" name="start_date" type="date" value="2026-04-20" required />
            </div>
            <div class="field">
              <label for="transfer_end_date">结束日期</label>
              <input id="transfer_end_date" name="end_date" type="date" value="2026-05-19" required />
            </div>
          </div>
          <div class="field">
            <label for="transfer_output_dir">输出区：更新后汇总表保存地址</label>
            <div class="path-row">
              <input id="transfer_output_dir" name="output_dir" type="text" value="D:/pycharm/xls-producer/web_work/outputs" required />
              <button id="chooseTransferOutputDir" class="secondary" type="button">选择文件夹</button>
            </div>
            <div class="hint">默认使用上传汇总表的原文件名保存到该目录；同名文件会被覆盖。</div>
          </div>
          <div class="actions">
            <button id="transferSubmit" type="submit">更新中转站情况</button>
          </div>
          <div id="transferStatus" class="status">等待上传中转站台账和待修改汇总表。</div>
          <div id="transferDownloads" class="download-list"></div>
        </form>
      </section>
      <aside>
        <h2 class="summary-title">生成结果</h2>
        <div class="kv">
          <span>时间跨度</span><strong id="rangeText">2026-04-20 至 2026-05-19</strong>
          <span>月份依据</span><strong>结束日期</strong>
          <span>输出月份</span><strong id="monthText">2026 年 5 月</strong>
          <span>保存位置</span><strong id="saveText">D:/pycharm/xls-producer/web_work/outputs</strong>
        </div>
        <div class="counts" id="counts">
          <span>暂无数据</span><strong>-</strong>
        </div>
        <ul class="missing" id="missing"></ul>
        <div class="notes">
          <p><strong>使用须知</strong></p>
          <p>先生成汇总表，再用汇总表和源台账生成两个 Word 报告。</p>
          <p>模板和台账请保持工作表名称、街道列和日期列结构稳定。</p>
          <p>如果目标文件已打开，生成时可能无法覆盖，请先关闭同名文件。</p>
        </div>
      </aside>
    </div>
    </div>
    </div>
    <div id="reportView" class="view" hidden>
      <div class="modal-panel" role="dialog" aria-modal="true">
      <div class="view-header">
        <h2>生成工作报告</h2>
        <button class="secondary back-home" type="button">返回入口</button>
      </div>
      <section>
        <h2 class="summary-title">生成两个 Word 报告</h2>
        <form id="reportForm">
          <div class="field">
            <label for="summary_xlsx">各街道问题汇总表</label>
            <input id="summary_xlsx" name="summary_xlsx" type="file" accept=".xlsx" required />
          </div>
          <div class="field">
            <label for="source_ledger">源台账</label>
            <input id="source_ledger" name="source_ledger" type="file" accept=".xlsx" required />
          </div>
          <div class="grid">
            <div class="field">
              <label for="residential_template">垃圾分类检查自查报告模板</label>
              <input id="residential_template" name="residential_template" type="file" accept=".docx" required />
            </div>
            <div class="field">
              <label for="social_template">社会单位、餐饮单位检查报告模板</label>
              <input id="social_template" name="social_template" type="file" accept=".docx" required />
            </div>
          </div>
          <div class="grid">
            <div class="field">
              <label for="report_start_date">开始日期</label>
              <input id="report_start_date" name="start_date" type="date" value="2026-04-20" required />
            </div>
            <div class="field">
              <label for="report_end_date">结束日期</label>
              <input id="report_end_date" name="end_date" type="date" value="2026-05-19" required />
            </div>
          </div>
          <div class="field">
            <label for="report_output_dir">报告保存地址</label>
            <div class="path-row">
              <input id="report_output_dir" name="output_dir" type="text" value="D:/pycharm/xls-producer/web_work/outputs" required />
              <button id="chooseReportOutputDir" class="secondary" type="button">选择文件夹</button>
            </div>
          </div>
          <div class="actions">
            <button id="reportSubmit" type="submit">生成两个报告</button>
          </div>
          <div id="reportStatus" class="status">等待上传报告模板。</div>
          <div id="reportDownloads" class="download-list"></div>
        </form>
      </section>
      </div>
    </div>
  </main>
  <script>
    const form = document.getElementById("form");
    const startInput = document.getElementById("start_date");
    const endInput = document.getElementById("end_date");
    const outputDirInput = document.getElementById("output_dir");
    const chooseOutputDir = document.getElementById("chooseOutputDir");
    const outputName = document.getElementById("outputName");
    const rangeText = document.getElementById("rangeText");
    const monthText = document.getElementById("monthText");
    const saveText = document.getElementById("saveText");
    const statusBox = document.getElementById("status");
    const submit = document.getElementById("submit");
    const download = document.getElementById("download");
    const counts = document.getElementById("counts");
    const missing = document.getElementById("missing");
    const reportForm = document.getElementById("reportForm");
    const reportSubmit = document.getElementById("reportSubmit");
    const reportStatus = document.getElementById("reportStatus");
    const reportDownloads = document.getElementById("reportDownloads");
    const reportStartInput = document.getElementById("report_start_date");
    const reportEndInput = document.getElementById("report_end_date");
    const reportOutputDirInput = document.getElementById("report_output_dir");
    const chooseReportOutputDir = document.getElementById("chooseReportOutputDir");
    const purityForm = document.getElementById("purityForm");
    const puritySubmit = document.getElementById("puritySubmit");
    const purityStatus = document.getElementById("purityStatus");
    const purityDownloads = document.getElementById("purityDownloads");
    const purityStartInput = document.getElementById("purity_start_date");
    const purityEndInput = document.getElementById("purity_end_date");
    const purityOutputDirInput = document.getElementById("purity_output_dir");
    const choosePurityOutputDir = document.getElementById("choosePurityOutputDir");
    const cityForm = document.getElementById("cityForm");
    const citySubmit = document.getElementById("citySubmit");
    const cityStatus = document.getElementById("cityStatus");
    const cityDownloads = document.getElementById("cityDownloads");
    const cityStartInput = document.getElementById("city_start_date");
    const cityEndInput = document.getElementById("city_end_date");
    const cityOutputDirInput = document.getElementById("city_output_dir");
    const chooseCityOutputDir = document.getElementById("chooseCityOutputDir");
    const transferForm = document.getElementById("transferForm");
    const transferSubmit = document.getElementById("transferSubmit");
    const transferStatus = document.getElementById("transferStatus");
    const transferDownloads = document.getElementById("transferDownloads");
    const transferStartInput = document.getElementById("transfer_start_date");
    const transferEndInput = document.getElementById("transfer_end_date");
    const transferOutputDirInput = document.getElementById("transfer_output_dir");
    const chooseTransferOutputDir = document.getElementById("chooseTransferOutputDir");
    const homeView = document.getElementById("homeView");
    const summaryView = document.getElementById("summaryView");
    const reportView = document.getElementById("reportView");
    const openSummary = document.getElementById("openSummary");
    const openReport = document.getElementById("openReport");

    function showView(viewName) {
      homeView.hidden = viewName !== "home";
      summaryView.hidden = viewName !== "summary";
      reportView.hidden = viewName !== "report";
      document.body.classList.toggle("modal-open", viewName !== "home");
      if (viewName === "home") {
        window.scrollTo({ top: 0, behavior: "smooth" });
      }
    }

    openSummary.addEventListener("click", () => showView("summary"));
    openReport.addEventListener("click", () => showView("report"));
    document.querySelectorAll(".back-home").forEach(button => {
      button.addEventListener("click", () => showView("home"));
    });
    document.querySelectorAll(".view").forEach(view => {
      view.addEventListener("click", event => {
        if (event.target === view) showView("home");
      });
    });
    document.addEventListener("keydown", event => {
      if (event.key === "Escape" && (!summaryView.hidden || !reportView.hidden)) {
        showView("home");
      }
    });

    function updatePreview() {
      const start = startInput.value;
      const end = endInput.value;
      if (!end) return;
      const date = new Date(end + "T00:00:00");
      const year = date.getFullYear();
      const month = date.getMonth() + 1;
      outputName.textContent = `（汇总）${year}.${month}月各街道问题汇总.xlsx`;
      rangeText.textContent = `${start || "-"} 至 ${end}`;
      monthText.textContent = `${year} 年 ${month} 月`;
      saveText.textContent = outputDirInput.value || "-";
    }

    function setStatus(text, cls) {
      statusBox.className = "status" + (cls ? " " + cls : "");
      statusBox.textContent = text;
    }

    startInput.addEventListener("change", updatePreview);
    endInput.addEventListener("change", updatePreview);
    outputDirInput.addEventListener("input", updatePreview);
    updatePreview();

    async function chooseDirectory(input, afterChange) {
      const current = encodeURIComponent(input.value || "");
      const response = await fetch(`/choose-output-dir?current=${current}`);
      const result = await response.json();
      if (!response.ok || !result.ok) {
        throw new Error(result.error || "选择文件夹失败");
      }
      if (result.path) {
        input.value = result.path;
        if (afterChange) afterChange();
      }
    }

    chooseOutputDir.addEventListener("click", async () => {
      chooseOutputDir.disabled = true;
      try {
        await chooseDirectory(outputDirInput, updatePreview);
      } catch (error) {
        setStatus(error.message, "error");
      } finally {
        chooseOutputDir.disabled = false;
      }
    });

    chooseReportOutputDir.addEventListener("click", async () => {
      chooseReportOutputDir.disabled = true;
      try {
        await chooseDirectory(reportOutputDirInput);
      } catch (error) {
        reportStatus.className = "status error";
        reportStatus.textContent = error.message;
      } finally {
        chooseReportOutputDir.disabled = false;
      }
    });

    choosePurityOutputDir.addEventListener("click", async () => {
      choosePurityOutputDir.disabled = true;
      try {
        await chooseDirectory(purityOutputDirInput);
      } catch (error) {
        purityStatus.className = "status error";
        purityStatus.textContent = error.message;
      } finally {
        choosePurityOutputDir.disabled = false;
      }
    });

    chooseCityOutputDir.addEventListener("click", async () => {
      chooseCityOutputDir.disabled = true;
      try {
        await chooseDirectory(cityOutputDirInput);
      } catch (error) {
        cityStatus.className = "status error";
        cityStatus.textContent = error.message;
      } finally {
        chooseCityOutputDir.disabled = false;
      }
    });

    chooseTransferOutputDir.addEventListener("click", async () => {
      chooseTransferOutputDir.disabled = true;
      try {
        await chooseDirectory(transferOutputDirInput);
      } catch (error) {
        transferStatus.className = "status error";
        transferStatus.textContent = error.message;
      } finally {
        chooseTransferOutputDir.disabled = false;
      }
    });

    form.addEventListener("submit", async (event) => {
      event.preventDefault();
      download.hidden = true;
      missing.innerHTML = "";

      if (startInput.value > endInput.value) {
        setStatus("开始日期不能晚于结束日期。", "error");
        return;
      }

      submit.disabled = true;
      setStatus("正在上传并生成，请稍候...", "");
      try {
        const body = new FormData(form);
        const response = await fetch("/generate", { method: "POST", body });
        const result = await response.json();
        if (!response.ok || !result.ok) {
          throw new Error(result.error || "生成失败");
        }

        const message = `生成完成：${result.filename}\n保存位置：${result.output_path}`;
        setStatus(message, result.missing.length ? "warn" : "ok");
        download.href = result.download_url;
        download.download = result.filename;
        download.hidden = false;

        counts.innerHTML = "";
        for (const [name, count] of Object.entries(result.row_counts)) {
          counts.insertAdjacentHTML("beforeend", `<span>${name}</span><strong>${count}</strong>`);
        }
        missing.innerHTML = result.missing.map(item => `<li>${item}</li>`).join("");
      } catch (error) {
        setStatus(error.message, "error");
      } finally {
        submit.disabled = false;
      }
    });

    reportForm.addEventListener("submit", async (event) => {
      event.preventDefault();
      reportDownloads.innerHTML = "";

      if (reportStartInput.value > reportEndInput.value) {
        reportStatus.className = "status error";
        reportStatus.textContent = "开始日期不能晚于结束日期。";
        return;
      }

      reportSubmit.disabled = true;
      reportStatus.className = "status";
      reportStatus.textContent = "正在上传并生成两个报告，请稍候...";
      try {
        const body = new FormData(reportForm);
        const response = await fetch("/generate-reports", { method: "POST", body });
        const result = await response.json();
        if (!response.ok || !result.ok) {
          throw new Error(result.error || "报告生成失败");
        }

        reportStatus.className = "status" + (result.missing.length ? " warn" : " ok");
        reportStatus.textContent = `生成完成：${result.files.length} 个报告\n保存位置：${result.output_dir}`;
        reportDownloads.innerHTML = result.files.map(file =>
          `<a class="download" href="${file.download_url}" download="${file.filename}">${file.filename}</a>`
        ).join("");
      } catch (error) {
        reportStatus.className = "status error";
        reportStatus.textContent = error.message;
      } finally {
        reportSubmit.disabled = false;
      }
    });

    purityForm.addEventListener("submit", async (event) => {
      event.preventDefault();
      purityDownloads.innerHTML = "";

      if (purityStartInput.value > purityEndInput.value) {
        purityStatus.className = "status error";
        purityStatus.textContent = "开始日期不能晚于结束日期。";
        return;
      }

      puritySubmit.disabled = true;
      purityStatus.className = "status";
      purityStatus.textContent = "正在生成对照表并更新汇总表，请稍候...";
      try {
        const body = new FormData(purityForm);
        const response = await fetch("/update-purity", { method: "POST", body });
        const result = await response.json();
        if (!response.ok || !result.ok) {
          throw new Error(result.error || "更新失败");
        }

        purityStatus.className = "status ok";
        purityStatus.textContent = `更新完成：${result.summary_filename}\n汇总表位置：${result.summary_path}\n对照表位置：${result.process_path}\n更新容器总数行：${result.updated.container_rows}，更新纯净率行：${result.updated.purity_rows}`;
        purityDownloads.innerHTML = [
          `<a class="download" href="${result.summary_download_url}" download="${result.summary_filename}">下载更新后汇总表</a>`,
          `<a class="download" href="${result.process_download_url}" download="${result.process_filename}">下载容器对照表</a>`
        ].join("");
      } catch (error) {
        purityStatus.className = "status error";
        purityStatus.textContent = error.message;
      } finally {
        puritySubmit.disabled = false;
      }
    });

    cityForm.addEventListener("submit", async (event) => {
      event.preventDefault();
      cityDownloads.innerHTML = "";

      if (cityStartInput.value > cityEndInput.value) {
        cityStatus.className = "status error";
        cityStatus.textContent = "开始日期不能晚于结束日期。";
        return;
      }

      citySubmit.disabled = true;
      cityStatus.className = "status";
      cityStatus.textContent = "正在读取市级汇总表并更新月汇总表，请稍候...";
      try {
        const body = new FormData(cityForm);
        const response = await fetch("/update-city", { method: "POST", body });
        const result = await response.json();
        if (!response.ok || !result.ok) {
          throw new Error(result.error || "更新失败");
        }

        cityStatus.className = "status ok";
        cityStatus.textContent = `更新完成：${result.summary_filename}\n汇总表位置：${result.summary_path}\n更新市级行：${result.updated.city_rows}\n小区表街道数：${result.resident_street_count}，社会表街道数：${result.social_street_count}\n时间范围：${result.start_date} 至 ${result.end_date}`;
        cityDownloads.innerHTML = `<a class="download" href="${result.summary_download_url}" download="${result.summary_filename}">下载更新后汇总表</a>`;
      } catch (error) {
        cityStatus.className = "status error";
        cityStatus.textContent = error.message;
      } finally {
        citySubmit.disabled = false;
      }
    });

    transferForm.addEventListener("submit", async (event) => {
      event.preventDefault();
      transferDownloads.innerHTML = "";

      if (transferStartInput.value > transferEndInput.value) {
        transferStatus.className = "status error";
        transferStatus.textContent = "开始日期不能晚于结束日期。";
        return;
      }

      transferSubmit.disabled = true;
      transferStatus.className = "status";
      transferStatus.textContent = "正在读取中转站台账并更新月汇总表，请稍候...";
      try {
        const body = new FormData(transferForm);
        const response = await fetch("/update-transfer", { method: "POST", body });
        const result = await response.json();
        if (!response.ok || !result.ok) {
          throw new Error(result.error || "更新失败");
        }

        transferStatus.className = "status ok";
        const diagnostics = result.diagnostics || {};
        const whiteSamples = (diagnostics.white_problem_samples || []).map((sample) => {
          const metrics = (sample.matched_metrics || []).join("、") || "无";
          const problems = (sample.problems || []).join("、") || "无";
          return `白纸坊样例：日期=${sample.date || ""} 范围内=${sample.in_range ? "是" : "否"} 问题=${problems} 命中=${metrics}`;
        }).join("\n");
        transferStatus.textContent = `更新完成：${result.summary_filename}\n汇总表位置：${result.summary_path}\n更新中转站行：${result.updated.transfer_rows}\n识别街道数：${result.street_count}\n时间范围：${result.start_date} 至 ${result.end_date}\n诊断：工作表=${diagnostics.sheet || ""}，表头行=${diagnostics.header_row || ""}，日期列=${diagnostics.date_column || ""}，街道列=${diagnostics.street_column || ""}，具体问题列=${(diagnostics.problem_text_columns || []).join("、") || "无"}，范围内行=${diagnostics.rows_in_date_range || 0}/${diagnostics.rows_seen || 0}，白纸坊范围内行=${diagnostics.white_rows_in_date_range || 0}/${diagnostics.white_rows_seen || 0}${whiteSamples ? "\n" + whiteSamples : ""}`;
        transferDownloads.innerHTML = `<a class="download" href="${result.summary_download_url}" download="${result.summary_filename}">下载更新后汇总表</a>`;
      } catch (error) {
        transferStatus.className = "status error";
        transferStatus.textContent = error.message;
      } finally {
        transferSubmit.disabled = false;
      }
    });
  </script>
</body>
</html>
"""


def json_response(handler: BaseHTTPRequestHandler, status: int, payload: dict) -> None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def save_upload(field, directory: Path, fallback_name: str, allowed_extensions: tuple[str, ...] = (".xlsx",)) -> Path:
    filename = Path(field.filename or fallback_name).name
    if not filename.lower().endswith(allowed_extensions):
        allowed = "、".join(allowed_extensions)
        raise ValueError(f"{filename} 不是 {allowed} 文件")
    stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    target = directory / f"{stamp}_{filename}"
    with target.open("wb") as output:
        shutil.copyfileobj(field.file, output)
    return target


def resolve_output_dir(value: str | None) -> Path:
    text = (value or "").strip().strip('"')
    path = Path(text) if text else DEFAULT_OUTPUT_DIR
    if not path.is_absolute():
        path = (BASE_DIR / path).resolve()
    path.mkdir(parents=True, exist_ok=True)
    if not path.is_dir():
        raise ValueError(f"{path} 不是有效文件夹")
    return path


def register_download(path: Path) -> str:
    key = f"{datetime.now():%Y%m%d%H%M%S%f}_{path.name}"
    DOWNLOAD_FILES[key] = path
    return f"/download/{quote(key)}"


def first_sheet_name(wb, keyword: str) -> str:
    for name in wb.sheetnames:
        if keyword in name:
            return name
    raise ValueError(f"未找到包含“{keyword}”的工作表")


def header_index(headers: list[str], keyword: str) -> int:
    for idx, header in enumerate(headers):
        if keyword in header:
            return idx
    raise ValueError(f"原始台账中未找到“{keyword}”列")


def optional_header_index(headers: list[str], keyword: str) -> int | None:
    for idx, header in enumerate(headers):
        if keyword in header:
            return idx
    return None


def excel_file_for_openpyxl(path: Path) -> tuple[Path, Path | None]:
    if path.suffix.lower() != ".xls":
        return path, None

    converted = PROCESS_DIR / f"{path.stem}_{datetime.now():%Y%m%d%H%M%S%f}_converted.xlsx"
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError("当前环境缺少 Excel 转换组件，无法读取 xls，请先另存为 xlsx 后上传。") from exc

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(path.resolve()))
        workbook.SaveAs(str(converted.resolve()), FileFormat=51)
    except Exception as exc:
        raise RuntimeError("xls 文件转换失败，请确认本机已安装 Excel，或先另存为 xlsx 后上传。") from exc
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()

    return converted, converted


CHINESE_DIGITS = {
    "零": 0,
    "一": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}


def parse_chinese_integer(text: str) -> int | None:
    if not text:
        return None
    if text == "十":
        return 10
    if "十" in text:
        left, _, right = text.partition("十")
        tens = CHINESE_DIGITS.get(left, 1 if left == "" else None)
        ones = CHINESE_DIGITS.get(right, 0 if right == "" else None)
        if tens is None or ones is None:
            return None
        return tens * 10 + ones
    return CHINESE_DIGITS.get(text)


def bucket_count_from_header(header: str) -> int | None:
    text = clean_text(header)
    match = re.search(r"(?:有)?(\d+)(?:个垃圾桶|组桶|个桶)", text)
    if match:
        return int(match.group(1))
    match = re.search(r"(?:有)?([零一二两三四五六七八九十]+)(?:个垃圾桶|组桶|个桶)", text)
    if match:
        return parse_chinese_integer(match.group(1))
    return None


def indicator_value(value) -> float:
    numeric = number(value)
    if numeric:
        return numeric
    text = clean_text(value)
    if not text or text in {"0", "否", "无", "不涉及", "nan", "None"}:
        return 0.0
    return 1.0


def combined_headers(ws, row1: int = 1, row2: int = 2) -> list[str]:
    headers = []
    for col in range(1, ws.max_column + 1):
        top = clean_text(ws.cell(row1, col).value)
        bottom = clean_text(ws.cell(row2, col).value)
        if top and bottom and top not in bottom:
            headers.append(f"{top}{bottom}")
        else:
            headers.append(bottom or top)
    return headers


def find_ledger_header_context(ws) -> tuple[int, list[str]]:
    for row in range(1, min(ws.max_row, 12) + 1):
        if row < ws.max_row:
            current = [clean_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
            next_row = [clean_text(ws.cell(row + 1, col).value) for col in range(1, ws.max_column + 1)]
            joined = []
            for top, bottom in zip(current, next_row):
                if top and bottom and top not in bottom:
                    joined.append(f"{top}{bottom}")
                else:
                    joined.append(bottom or top)
            if any("街道" in header or "3级点位" in header for header in next_row):
                return row + 1, joined
        current = [clean_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        previous = [clean_text(ws.cell(row - 1, col).value) if row > 1 else "" for col in range(1, ws.max_column + 1)]
        joined = []
        for top, bottom in zip(previous, current):
            if top and bottom and top not in bottom:
                joined.append(f"{top}{bottom}")
            else:
                joined.append(bottom or top)
        if any("街道" in header or "3级点位" in header for header in joined):
            return row, joined
    return 2, combined_headers(ws, 1, 2)


def parse_date_like(value) -> datetime | None:
    parsed = to_datetime(value)
    if parsed is not None:
        return parsed
    text = clean_text(value)
    for fmt in ("%Y-%m-%d%H:%M:%S", "%Y-%m-%d%H:%M", "%Y/%m/%d%H:%M:%S", "%Y/%m/%d%H:%M"):
        try:
            parsed = datetime.strptime(text, fmt)
            if parsed.year == 1016:
                parsed = parsed.replace(year=2026)
            return parsed
        except ValueError:
            pass
    match = re.match(r"(1016|2026)([-/.]\d{1,2}[-/.]\d{1,2})(?:\s*|T)?(\d{1,2}:\d{2}(?::\d{2})?)?", text)
    if match:
        normalized = f"2026{match.group(2).replace('/', '-').replace('.', '-')}"
        if match.group(3):
            normalized = f"{normalized} {match.group(3)}"
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(normalized, fmt)
            except ValueError:
                pass
    match = re.match(r"(\d{8})", text)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y%m%d")
        except ValueError:
            return None
    return None


def build_container_comparison(
    source_file: Path,
    start_date: datetime,
    end_date: datetime,
    process_file: Path,
    bad_counts: dict[str, float] | None = None,
) -> dict[str, dict[str, float]]:
    readable_file, temp_file = excel_file_for_openpyxl(source_file)
    wb = load_workbook(readable_file, read_only=True, data_only=True)
    detail_rows: list[list] = []
    try:
        try:
            ws = wb[first_sheet_name(wb, "居民")]
        except ValueError:
            ws = wb[wb.sheetnames[0]]
        header_row, headers = find_ledger_header_context(ws)
        plain_headers = [clean_text(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
        street_idx = optional_header_index(headers, "街道")
        if street_idx is None:
            street_idx = optional_header_index(headers, "3级点位")
        if street_idx is None:
            street_idx = header_index(plain_headers, "街道")
        date_idx = optional_header_index(headers, "日期")
        if date_idx is None:
            date_idx = optional_header_index(headers, "时间")
        if date_idx is None:
            date_idx = optional_header_index(headers, "编号")
        if date_idx is None:
            date_idx = 0
        container_idx = optional_header_index(headers, "容器总数")
        bucket_columns = [
            (idx, bucket_count, header)
            for idx, header in enumerate(headers)
            if (bucket_count := bucket_count_from_header(header)) is not None
        ]

        data: dict[str, dict[str, float]] = {}
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            checked_at = parse_date_like(row[date_idx] if date_idx < len(row) else None)
            if checked_at and not (start_date <= checked_at <= end_date):
                continue
            street = street_key(row[street_idx] if street_idx < len(row) else None)
            if not street:
                continue
            current = data.setdefault(street, {"container_total": 0.0, "bad_total": 0.0, "purity_rate": 1.0})
            row_bucket_added = 0.0
            if bucket_columns:
                for idx, bucket_count, header in bucket_columns:
                    times = indicator_value(row[idx] if idx < len(row) else None)
                    if times <= 0:
                        continue
                    added = bucket_count * times
                    current["container_total"] += added
                    row_bucket_added += added
                    detail_rows.append([street, row_number, header, bucket_count, times, added])
            for idx, value in enumerate(row):
                bucket_count = bucket_count_from_header(clean_text(value))
                if bucket_count is None:
                    continue
                current["container_total"] += bucket_count
                row_bucket_added += bucket_count
                detail_rows.append([street, row_number, headers[idx] if idx < len(headers) else f"第{idx + 1}列", bucket_count, 1, bucket_count])
            if row_bucket_added <= 0 and container_idx is not None:
                current["container_total"] += number(row[container_idx] if container_idx < len(row) else None)

        for street, bad_count in (bad_counts or {}).items():
            current = data.setdefault(street, {"container_total": 0.0, "bad_total": 0.0, "purity_rate": 1.0})
            current["bad_total"] = bad_count

        for values in data.values():
            containers = values["container_total"]
            bad = values["bad_total"]
            values["purity_rate"] = 1 - bad / containers if containers else 1
    finally:
        wb.close()
        if temp_file is not None:
            temp_file.unlink(missing_ok=True)

    process_file.parent.mkdir(parents=True, exist_ok=True)
    out = Workbook()
    out_ws = out.active
    out_ws.title = "容器总数与纯净率对照表"
    out_ws.append(["街道名称", "识别桶数", "桶内分类不纯净", "桶内分类纯净率"])
    for street in sorted(data):
        values = data[street]
        out_ws.append([
            street,
            int(values["container_total"]) if values["container_total"].is_integer() else values["container_total"],
            int(values["bad_total"]) if values["bad_total"].is_integer() else values["bad_total"],
            round(values["purity_rate"], 4),
        ])
    detail_ws = out.create_sheet("桶数识别明细")
    detail_ws.append(["街道名称", "源表行号", "命中问题列", "单项桶数", "记录次数", "计入桶数"])
    for detail in detail_rows:
        detail_ws.append(detail)
    out.save(process_file)
    return data


def update_summary_container_and_purity(summary_file: Path, output_file: Path, comparison: dict[str, dict[str, float]]) -> dict[str, int]:
    wb = load_workbook(summary_file)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    street_columns = [
        (col, street_key(ws.cell(1, col).value))
        for col in range(4, ws.max_column + 1)
        if street_key(ws.cell(1, col).value)
    ]
    updated = {"container_rows": 0, "purity_rows": 0}
    current_category = ""
    for row in range(2, ws.max_row + 1):
        category = clean_text(ws.cell(row, 2).value)
        metric = clean_text(ws.cell(row, 3).value)
        if category:
            current_category = category
        row_text = metric_key(f"{current_category}{metric}")
        is_container_row = "容器总数" in row_text
        is_purity_row = "桶内分类纯净率" in row_text or (metric_key(current_category) == "纯净率" and "纯净率" in row_text)
        if not is_container_row and not is_purity_row:
            continue
        for col, street in street_columns:
            values = comparison.get(street)
            if not values:
                continue
            if is_container_row:
                total = values["container_total"]
                ws.cell(row, col).value = int(total) if total.is_integer() else total
            elif is_purity_row:
                ws.cell(row, col).value = round(values["purity_rate"], 4)
        if is_container_row:
            updated["container_rows"] += 1
        if is_purity_row:
            updated["purity_rows"] += 1

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()
    return updated


def read_summary_bad_counts(summary_file: Path) -> dict[str, float]:
    wb = load_workbook(summary_file, read_only=True, data_only=True)
    try:
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
        street_columns = [
            (col, street_key(ws.cell(1, col).value))
            for col in range(4, ws.max_column + 1)
            if street_key(ws.cell(1, col).value)
        ]
        current_category = ""
        for row in range(2, ws.max_row + 1):
            category = clean_text(ws.cell(row, 2).value)
            metric = clean_text(ws.cell(row, 3).value)
            if category:
                current_category = category
            if current_category == "分类设施管理达标情况" and metric_key(metric) == "桶内分类不纯净":
                return {
                    street: number(ws.cell(row, col).value)
                    for col, street in street_columns
                }
    finally:
        wb.close()
    raise ValueError("月汇总表中未找到“分类设施管理达标情况 / 桶内分类不纯净”行")


def find_header_row(ws, keywords: tuple[str, ...], max_scan_rows: int = 20) -> tuple[int, list[str]]:
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = [clean_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if all(any(keyword in header for header in headers) for keyword in keywords):
            return row, headers
    raise ValueError(f"{ws.title} 未找到包含“{'、'.join(keywords)}”的表头行")


def find_optional_city_date_index(headers: list[str]) -> int | None:
    for keyword in ("检查日期", "日期", "时间", "编号"):
        idx = optional_header_index(headers, keyword)
        if idx is not None:
            return idx
    return None


def find_city_street_index(headers: list[str]) -> int:
    for keyword in ("属地街道", "所属街道", "街道名称", "街道"):
        idx = optional_header_index(headers, keyword)
        if idx is not None:
            return idx
    raise ValueError("市级汇总表中未找到街道列")


def find_city_header_row(ws, max_scan_rows: int = 30) -> tuple[int, list[str]]:
    best: tuple[float, int, list[str]] | None = None
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = [clean_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        date_idx = find_optional_city_date_index(headers)
        try:
            street_idx = find_city_street_index(headers)
        except ValueError:
            street_idx = None
        if date_idx is None or street_idx is None:
            continue
        problem_idx = optional_header_index(headers, "问题总数")
        score = 10
        if date_idx < 20:
            score += 10
        if street_idx < 20:
            score += 10
        if problem_idx is not None and problem_idx < 20:
            score += 5
        if optional_header_index(headers, "序号") is not None:
            score += 2
        nonempty = sum(1 for header in headers[:20] if header)
        score += min(nonempty, 10) / 10
        candidate = (score, row, headers)
        if best is None or candidate[0] > best[0]:
            best = candidate
    if best is None:
        raise ValueError(f"{ws.title} 未找到包含“日期、街道”的表头行")
    return best[1], best[2]


def read_city_summary_stats(summary_file: Path, start_date: datetime, end_date: datetime) -> dict[str, dict[str, float]]:
    wb = load_workbook(summary_file, read_only=True, data_only=True)
    stats: dict[str, dict[str, float]] = {}
    try:
        for ws in wb.worksheets:
            try:
                header_row, headers = find_city_header_row(ws)
            except ValueError:
                continue
            street_idx = find_city_street_index(headers)
            problem_idx = optional_header_index(headers, "问题总数")
            date_idx = find_optional_city_date_index(headers)
            if date_idx is None:
                continue
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                checked_at = parse_date_like(row[date_idx] if date_idx < len(row) else None)
                if checked_at is None or not (start_date <= checked_at <= end_date):
                    continue
                street = street_key(row[street_idx] if street_idx < len(row) else None)
                if not street:
                    continue
                current = stats.setdefault(street, {"record_count": 0.0, "problem_total": 0.0})
                current["record_count"] += 1
                if problem_idx is not None:
                    current["problem_total"] += number(row[problem_idx] if problem_idx < len(row) else None)
            break
    finally:
        wb.close()
    if not stats:
        raise ValueError(f"{summary_file.name} 在所选时间段内未解析到街道数据")
    return stats


def update_summary_city_checks(
    summary_file: Path,
    output_file: Path,
    resident_stats: dict[str, dict[str, float]],
    social_stats: dict[str, dict[str, float]],
) -> dict[str, int]:
    wb = load_workbook(summary_file)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    street_columns = [
        (col, street_key(ws.cell(1, col).value))
        for col in range(4, ws.max_column + 1)
        if street_key(ws.cell(1, col).value)
    ]
    updated = {"city_rows": 0}
    current_category = ""
    for row in range(2, ws.max_row + 1):
        category = clean_text(ws.cell(row, 2).value)
        metric = clean_text(ws.cell(row, 3).value)
        if category:
            current_category = category
        if current_category != "市级检查情况":
            continue
        metric = metric_key(metric)
        if metric not in ("检查小区数", "检查社会单位数", "市级检查问题数"):
            continue
        for col, street in street_columns:
            resident = resident_stats.get(street, {"record_count": 0.0, "problem_total": 0.0})
            social = social_stats.get(street, {"record_count": 0.0, "problem_total": 0.0})
            if metric == "检查小区数":
                value = resident["record_count"]
            elif metric == "检查社会单位数":
                value = social["record_count"]
            else:
                value = resident["problem_total"]
            ws.cell(row, col).value = int(value) if float(value).is_integer() else value
        updated["city_rows"] += 1

    if updated["city_rows"] == 0:
        wb.close()
        raise ValueError("月汇总表中未找到“市级检查情况”相关行")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()
    return updated


def find_transfer_sheet(wb):
    for ws in wb.worksheets:
        if "中转" in ws.title:
            return ws
    if len(wb.worksheets) == 1:
        return wb.worksheets[0]
    raise ValueError("中转站台账中未找到包含“中转”的工作表")


def find_transfer_header_row(ws, max_scan_rows: int = 20) -> tuple[int, list[str]]:
    def is_header(headers: list[str]) -> bool:
        has_date = any("日期" in header or "时间" in header or "编号" in header for header in headers)
        has_street = any("街道" in header or "三级点位" in header or "3级点位" in header for header in headers)
        return has_date and has_street

    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = [clean_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if is_header(headers):
            return row, headers
        if row > 1:
            previous = [clean_text(ws.cell(row - 1, col).value) for col in range(1, ws.max_column + 1)]
            merged = []
            for top, bottom in zip(previous, headers):
                if top and bottom and top not in bottom:
                    merged.append(f"{top}{bottom}")
                else:
                    merged.append(bottom or top)
            if is_header(merged):
                return row, merged
    raise ValueError(f"{ws.title} 未找到包含“编号/日期、三级点位/街道”的表头行")


def find_transfer_date_index(headers: list[str]) -> int:
    for keyword in ("检查日期", "日期", "时间", "编号"):
        idx = optional_header_index(headers, keyword)
        if idx is not None:
            return idx
    return 0


def find_transfer_street_index(headers: list[str]) -> int:
    for keyword in ("街道名称", "所属街道", "属地街道", "街道", "三级点位", "3级点位"):
        idx = optional_header_index(headers, keyword)
        if idx is not None:
            return idx
    raise ValueError("中转站台账中未找到街道列或三级点位列")


def parse_transfer_date(value) -> datetime | None:
    parsed = parse_date_like(value)
    if parsed is not None:
        return parsed
    text = clean_text(value)
    match = re.search(r"(20\d{2}|10\d{2})[-/.年]?(\d{1,2})[-/.月]?(\d{1,2})", text)
    if match:
        year = 2026 if match.group(1) == "1016" else int(match.group(1))
        try:
            return datetime(year, int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    match = re.search(r"(?<!\d)(\d{8})(?!\d)", text)
    if match:
        try:
            parsed = datetime.strptime(match.group(1), "%Y%m%d")
            if parsed.year == 1016:
                parsed = parsed.replace(year=2026)
            return parsed
        except ValueError:
            return None
    return None


def transfer_street_key(value) -> str:
    text = clean_text(value)
    if not text:
        return ""
    known = [
        "西长安街", "新街口", "月坛", "展览路", "德胜", "金融街", "什刹海",
        "大栅栏", "天桥", "椿树", "陶然亭", "广内", "牛街", "白纸坊",
        "广外", "广安门外", "广安门内", "广安门",
    ]
    for name in sorted(known, key=len, reverse=True):
        if name in text:
            return street_key(name)
    match = re.search(r"([\u4e00-\u9fa5]{1,12}?街道)", text)
    if match:
        return street_key(match.group(1))
    return street_key(text)


def transfer_metric_alias(metric: str) -> str:
    key = metric_key(metric)
    aliases = {
        metric_key("消防水源不合格"): metric_key("无消防安全水源"),
        metric_key("灭火器过期"): metric_key("灭火器不合格"),
        metric_key("无灭蝇措施"): metric_key("无灭蝇措施"),
        metric_key("灭蝇措施"): metric_key("无灭蝇措施"),
        metric_key("灭蚊蝇设施"): metric_key("无灭蝇措施"),
        metric_key("无灭蚊蝇设施"): metric_key("无灭蝇措施"),
        metric_key("七禁收八不准"): metric_key("无七禁收八不准承诺书"),
        metric_key("无七禁收八不准"): metric_key("无七禁收八不准承诺书"),
        metric_key("堆放混乱"): metric_key("未按规定区域存放物品"),
        metric_key("物品堆放混乱"): metric_key("未按规定区域存放物品"),
        metric_key("计量称不能使用"): metric_key("称重系统损坏"),
        metric_key("计量秤不能使用"): metric_key("称重系统损坏"),
        metric_key("称重计量不能使用"): metric_key("称重系统损坏"),
        metric_key("未配备电源箱"): metric_key("配电箱处堆放杂物"),
        metric_key("未配备配电箱"): metric_key("配电箱处堆放杂物"),
    }
    return aliases.get(key, key)


def transfer_metric_keyword_groups(metric: str) -> list[tuple[str, ...]]:
    key = transfer_metric_alias(metric)
    return transfer_metric_keyword_rules().get(key, [])


def transfer_metric_keyword_rules() -> dict[str, list[tuple[str, ...]]]:
    return {
        metric_key("无安全风险公告"): [
            ("安全风险公告",),
            ("风险公告",),
        ],
        metric_key("无灭蝇措施"): [
            ("灭蝇",),
            ("灭蚊蝇",),
            ("蚊蝇", "设施"),
            ("蚊蝇", "措施"),
        ],
        metric_key("无七禁收八不准承诺书"): [
            ("七禁收八不准",),
            ("七禁", "八不准"),
            ("八不准", "承诺"),
        ],
        metric_key("未按规定区域存放物品"): [
            ("未按规定", "存放"),
            ("规定区域", "存放"),
            ("堆放混乱",),
            ("物品", "堆放"),
        ],
        metric_key("配电箱处堆放杂物"): [
            ("配电箱", "堆放"),
            ("电源箱",),
            ("配电箱",),
        ],
        metric_key("称重系统损坏"): [
            ("称重", "损坏"),
            ("称重", "不能使用"),
            ("计量称", "不能使用"),
            ("计量秤", "不能使用"),
            ("计量", "不能使用"),
        ],
        metric_key("无消防安全水源"): [
            ("消防", "水源"),
        ],
        metric_key("灭火器不合格"): [
            ("灭火器", "不合格"),
            ("灭火器", "过期"),
            ("灭火器", "欠压"),
        ],
        metric_key("无备案公示"): [
            ("备案公示",),
            ("无备案",),
        ],
        metric_key("周边环境脏乱"): [
            ("周边", "脏乱"),
            ("周边", "不洁"),
            ("环境", "脏乱"),
        ],
        metric_key("清运不及时、可回收物大量积存"): [
            ("清运", "不及时"),
            ("可回收物", "积存"),
        ],
    }


def transfer_known_metric_keys() -> set[str]:
    return {
        metric_key("清运不及时、可回收物大量积存"),
        metric_key("周边环境脏乱"),
        metric_key("无可回收价格表"),
        metric_key("无备案公示"),
        metric_key("无消防安全水源"),
        metric_key("无营业执照"),
        metric_key("无安全风险公告"),
        metric_key("称重系统损坏"),
        metric_key("灭火器不合格"),
        metric_key("未按规定区域存放物品"),
        metric_key("无七禁收八不准承诺书"),
        metric_key("配电箱处堆放杂物"),
        metric_key("安全员未按时上岗"),
        metric_key("安全员无明显身份标识"),
        metric_key("未按时开门运行"),
        metric_key("无企安安"),
        metric_key("无灭蝇措施"),
        *transfer_metric_keyword_rules().keys(),
    }


def is_transfer_indicator_header(header: str) -> bool:
    key = metric_key(header)
    if not key:
        return False
    skip_keywords = (
        "检查日期", "日期", "时间", "街道", "点位", "名称", "运营主体", "具体问题", "问题描述",
        "问题情况", "问题合计", "合计", "备注", "编号", "序号", "定位", "照片", "图片",
        "指标", "检查员", "二维码",
    )
    return not any(keyword in key for keyword in skip_keywords)


def transfer_problem_value(value) -> float:
    text = clean_text(value)
    if text == "1":
        return 0.0
    numeric = number(value)
    if numeric:
        return numeric
    if not text or text in {"0", "否", "无", "不涉及", "nan", "None"}:
        return 0.0
    return 1.0


def find_transfer_problem_text_indexes(headers: list[str]) -> list[int]:
    keywords = ("具体问题", "问题描述", "问题情况", "存在问题", "检查问题")
    return [
        idx for idx, header in enumerate(headers)
        if any(keyword in clean_text(header) for keyword in keywords)
    ]


def split_transfer_problem_text(value) -> list[str]:
    text = clean_text(value)
    empty_values = {"0", "1", "否", "无", "不涉及", "nan", "None", "无问题", "未发现问题"}
    if not text or text in empty_values:
        return []
    parts = re.split(r"[;；,，、。.\n\r\t]+", text)
    results = []
    for part in parts:
        item = re.sub(r"^[（(]?\d+[）)、.．]?", "", part.strip())
        item = metric_key(item)
        if item and item not in empty_values:
            results.append(item)
    return results or [metric_key(text)]


def transfer_problem_matches(summary_metric: str, problem_text: str) -> bool:
    metric = transfer_metric_alias(summary_metric)
    problem = transfer_metric_alias(problem_text)
    if not metric or not problem:
        return False
    if metric in problem or problem in metric:
        return True
    problem_key = metric_key(problem_text)
    for group in transfer_metric_keyword_groups(metric):
        if all(keyword in problem_key for keyword in group):
            return True
    return False


def transfer_problem_metric_keys(problem_text: str) -> set[str]:
    problem = transfer_metric_alias(problem_text)
    known_metrics = transfer_known_metric_keys()
    matched: set[str] = set()
    if problem in known_metrics:
        matched.add(problem)
    for metric in known_metrics:
        if transfer_problem_matches(metric, problem_text):
            matched.add(metric)
    return matched


def read_transfer_station_stats(ledger_file: Path, start_date: datetime, end_date: datetime) -> dict[str, dict[str, float]]:
    readable_file, temp_file = excel_file_for_openpyxl(ledger_file)
    wb = load_workbook(readable_file, read_only=True, data_only=True)
    stats: dict[str, dict[str, float]] = {}
    diagnostics: dict[str, object] = {}
    try:
        ws = find_transfer_sheet(wb)
        header_row, headers = find_transfer_header_row(ws)
        date_idx = find_transfer_date_index(headers)
        street_idx = find_transfer_street_index(headers)
        problem_text_indexes = find_transfer_problem_text_indexes(headers)
        indicator_indexes = [
            idx for idx, header in enumerate(headers)
            if idx not in (date_idx, street_idx) and is_transfer_indicator_header(header)
        ]
        if not problem_text_indexes and not indicator_indexes:
            raise ValueError(f"{ws.title} 未找到“具体问题”列或可统计的问题列")

        diagnostics = {
            "sheet": ws.title,
            "header_row": header_row,
            "date_column": headers[date_idx] if date_idx < len(headers) else "",
            "street_column": headers[street_idx] if street_idx < len(headers) else "",
            "problem_text_columns": [headers[idx] for idx in problem_text_indexes],
            "indicator_columns": [headers[idx] for idx in indicator_indexes[:30]],
            "rows_seen": 0,
            "rows_in_date_range": 0,
            "rows_skipped_by_date": 0,
            "rows_skipped_by_street": 0,
            "white_rows_seen": 0,
            "white_rows_in_date_range": 0,
            "white_problem_samples": [],
        }

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            diagnostics["rows_seen"] = int(diagnostics["rows_seen"]) + 1
            checked_at = parse_transfer_date(row[date_idx] if date_idx < len(row) else None)
            raw_street = row[street_idx] if street_idx < len(row) else None
            street = transfer_street_key(raw_street)
            problem_items: list[str] = []
            for idx in problem_text_indexes:
                problem_items.extend(split_transfer_problem_text(row[idx] if idx < len(row) else None))
            if street == street_key("白纸坊"):
                diagnostics["white_rows_seen"] = int(diagnostics["white_rows_seen"]) + 1
                samples = diagnostics["white_problem_samples"]
                if isinstance(samples, list) and len(samples) < 8 and problem_items:
                    samples.append({
                        "date": checked_at.strftime("%Y-%m-%d") if checked_at else clean_text(row[date_idx] if date_idx < len(row) else None),
                        "in_range": bool(checked_at is not None and start_date <= checked_at <= end_date),
                        "street": clean_text(raw_street),
                        "problems": problem_items[:12],
                        "matched_metrics": sorted({
                            metric
                            for problem in problem_items
                            for metric in transfer_problem_metric_keys(problem)
                        }),
                    })
            if checked_at is None or not (start_date <= checked_at <= end_date):
                diagnostics["rows_skipped_by_date"] = int(diagnostics["rows_skipped_by_date"]) + 1
                continue
            diagnostics["rows_in_date_range"] = int(diagnostics["rows_in_date_range"]) + 1
            if not street:
                diagnostics["rows_skipped_by_street"] = int(diagnostics["rows_skipped_by_street"]) + 1
                continue
            if street == street_key("白纸坊"):
                diagnostics["white_rows_in_date_range"] = int(diagnostics["white_rows_in_date_range"]) + 1
            current = stats.setdefault(street, {"__record_count__": 0.0, "__issue_total__": 0.0, "__problem_texts__": []})
            current["__record_count__"] += 1
            if problem_items:
                current["__problem_texts__"].extend(problem_items)
                current["__issue_total__"] += len(problem_items)
                for problem in problem_items:
                    for metric in transfer_problem_metric_keys(problem):
                        current[metric] = current.get(metric, 0.0) + 1
            for idx in indicator_indexes:
                value = transfer_problem_value(row[idx] if idx < len(row) else None)
                if value == 0:
                    continue
                metric = transfer_metric_alias(headers[idx])
                current[metric] = current.get(metric, 0.0) + value
                current["__issue_total__"] += value
    finally:
        wb.close()
        if temp_file is not None:
            temp_file.unlink(missing_ok=True)

    if not stats:
        raise ValueError(f"{ledger_file.name} 在所选时间段内未解析到中转站数据")
    stats["__diagnostics__"] = diagnostics
    return stats


def update_summary_transfer_station(summary_file: Path, output_file: Path, transfer_stats: dict[str, dict[str, float]]) -> dict[str, int]:
    wb = load_workbook(summary_file)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    street_columns = [
        (col, street_key(ws.cell(1, col).value))
        for col in range(4, ws.max_column + 1)
        if street_key(ws.cell(1, col).value)
    ]
    updated = {"transfer_rows": 0}
    current_category = ""
    for row in range(2, ws.max_row + 1):
        category = clean_text(ws.cell(row, 2).value)
        metric = clean_text(ws.cell(row, 3).value)
        if category:
            current_category = category
        if current_category != "中转站" or not metric:
            continue
        metric = transfer_metric_alias(metric)
        for col, street in street_columns:
            values = transfer_stats.get(street, {})
            value = values.get(metric, 0)
            if not value:
                problem_texts = values.get("__problem_texts__", [])
                value = sum(1 for problem in problem_texts if transfer_problem_matches(metric, problem))
            ws.cell(row, col).value = value
        updated["transfer_rows"] += 1

    if updated["transfer_rows"] == 0:
        wb.close()
        raise ValueError("月汇总表中未找到“中转站”相关行")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()
    return updated


def choose_output_directory(current: str | None = None) -> str:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception as exc:
        raise RuntimeError("当前环境无法打开文件夹选择窗口，请手动填写保存路径。") from exc

    initial = (current or "").strip().strip('"')
    if not initial or not Path(initial).exists():
        initial = str(DEFAULT_OUTPUT_DIR)

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected = filedialog.askdirectory(
            title="选择保存文件夹",
            initialdir=initial,
            mustexist=False,
            parent=root,
        )
    finally:
        root.destroy()
    return selected


class AppHandler(BaseHTTPRequestHandler):
    server_version = "XlsProducerUI/2.0"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            body = INDEX_HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if parsed.path == "/choose-output-dir":
            self.choose_output_dir(parsed)
            return
        if parsed.path.startswith("/download/"):
            self.serve_download(parsed.path)
            return
        self.send_error(404)

    def do_POST(self) -> None:
        path = urlparse(self.path).path
        if path == "/update-purity":
            self.update_purity()
            return
        if path == "/update-city":
            self.update_city_checks()
            return
        if path == "/update-transfer":
            self.update_transfer_station()
            return
        if path == "/generate-reports":
            self.generate_report_docs()
            return
        if path != "/generate":
            self.send_error(404)
            return
        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            start_date = parse_ui_date(form.getfirst("start_date", ""))
            end_date = parse_ui_date(form.getfirst("end_date", ""), end_of_day=True)
            if start_date > end_date:
                raise ValueError("开始日期不能晚于结束日期")

            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            output_dir = resolve_output_dir(form.getfirst("output_dir", ""))
            template_file = save_upload(form["template"], UPLOAD_DIR, "template.xlsx")
            source_file = save_upload(form["source"], UPLOAD_DIR, "source.xlsx")
            filename = output_name_for(end_date)
            output_file = output_dir / filename

            result = generate_summary(source_file, template_file, start_date, end_date, output_file)
            json_response(
                self,
                200,
                {
                    "ok": True,
                    "filename": filename,
                    "download_url": register_download(output_file),
                    "output_path": str(output_file),
                    "row_counts": result["row_counts"],
                    "missing": result["missing"],
                },
            )
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def update_purity(self) -> None:
        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            start_date = parse_ui_date(form.getfirst("start_date", ""))
            end_date = parse_ui_date(form.getfirst("end_date", ""), end_of_day=True)
            if start_date > end_date:
                raise ValueError("开始日期不能晚于结束日期")

            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            PROCESS_DIR.mkdir(parents=True, exist_ok=True)
            output_dir = resolve_output_dir(form.getfirst("output_dir", ""))

            source_file = save_upload(form["source"], UPLOAD_DIR, "container_source.xlsx", (".xlsx", ".xls"))
            summary_original_name = Path(form["summary"].filename or "summary.xlsx").name
            summary_file = save_upload(form["summary"], UPLOAD_DIR, "summary.xlsx")
            process_file = PROCESS_DIR / "容器总数与纯净率对照表.xlsx"
            output_file = output_dir / summary_original_name

            bad_counts = read_summary_bad_counts(summary_file)
            comparison = build_container_comparison(source_file, start_date, end_date, process_file, bad_counts)
            updated = update_summary_container_and_purity(summary_file, output_file, comparison)
            json_response(
                self,
                200,
                {
                    "ok": True,
                    "summary_filename": output_file.name,
                    "summary_path": str(output_file),
                    "summary_download_url": register_download(output_file),
                    "process_filename": process_file.name,
                    "process_path": str(process_file),
                    "process_download_url": register_download(process_file),
                    "updated": updated,
                    "street_count": len(comparison),
                },
            )
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def update_city_checks(self) -> None:
        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            start_date = parse_ui_date(form.getfirst("start_date", ""))
            end_date = parse_ui_date(form.getfirst("end_date", ""), end_of_day=True)
            if start_date > end_date:
                raise ValueError("开始日期不能晚于结束日期")

            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            output_dir = resolve_output_dir(form.getfirst("output_dir", ""))

            summary_original_name = Path(form["summary"].filename or "summary.xlsx").name
            summary_file = save_upload(form["summary"], UPLOAD_DIR, "summary.xlsx")
            resident_file = save_upload(form["resident_summary"], UPLOAD_DIR, "city_resident.xlsx")
            social_file = save_upload(form["social_summary"], UPLOAD_DIR, "city_social.xlsx")
            output_file = output_dir / summary_original_name

            resident_stats = read_city_summary_stats(resident_file, start_date, end_date)
            social_stats = read_city_summary_stats(social_file, start_date, end_date)
            updated = update_summary_city_checks(summary_file, output_file, resident_stats, social_stats)
            json_response(
                self,
                200,
                {
                    "ok": True,
                    "summary_filename": output_file.name,
                    "summary_path": str(output_file),
                    "summary_download_url": register_download(output_file),
                    "updated": updated,
                    "resident_street_count": len(resident_stats),
                    "social_street_count": len(social_stats),
                    "start_date": start_date.strftime("%Y-%m-%d"),
                    "end_date": end_date.strftime("%Y-%m-%d"),
                },
            )
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def update_transfer_station(self) -> None:
        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            start_date = parse_ui_date(form.getfirst("start_date", ""))
            end_date = parse_ui_date(form.getfirst("end_date", ""), end_of_day=True)
            if start_date > end_date:
                raise ValueError("开始日期不能晚于结束日期")

            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            output_dir = resolve_output_dir(form.getfirst("output_dir", ""))

            summary_original_name = Path(form["summary"].filename or "summary.xlsx").name
            summary_file = save_upload(form["summary"], UPLOAD_DIR, "summary.xlsx")
            ledger_file = save_upload(form["transfer_ledger"], UPLOAD_DIR, "transfer_ledger.xlsx", (".xlsx", ".xls"))
            output_file = output_dir / summary_original_name

            transfer_stats = read_transfer_station_stats(ledger_file, start_date, end_date)
            updated = update_summary_transfer_station(summary_file, output_file, transfer_stats)
            diagnostics = transfer_stats.get("__diagnostics__", {})
            street_count = len([key for key in transfer_stats if not str(key).startswith("__")])
            json_response(
                self,
                200,
                {
                    "ok": True,
                    "summary_filename": output_file.name,
                    "summary_path": str(output_file),
                    "summary_download_url": register_download(output_file),
                    "updated": updated,
                    "street_count": street_count,
                    "diagnostics": diagnostics,
                    "start_date": start_date.strftime("%Y-%m-%d"),
                    "end_date": end_date.strftime("%Y-%m-%d"),
                },
            )
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def generate_report_docs(self) -> None:
        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            start_date = parse_ui_date(form.getfirst("start_date", ""))
            end_date = parse_ui_date(form.getfirst("end_date", ""), end_of_day=True)
            if start_date > end_date:
                raise ValueError("开始日期不能晚于结束日期")

            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            output_dir = resolve_output_dir(form.getfirst("output_dir", ""))
            summary_file = save_upload(form["summary_xlsx"], UPLOAD_DIR, "summary.xlsx")
            source_file = save_upload(form["source_ledger"], UPLOAD_DIR, "source.xlsx")
            residential_template = save_upload(
                form["residential_template"],
                UPLOAD_DIR,
                "residential_template.docx",
                (".docx",),
            )
            social_template = save_upload(
                form["social_template"],
                UPLOAD_DIR,
                "social_template.docx",
                (".docx",),
            )

            result = generate_reports(
                summary_file,
                source_file,
                residential_template,
                social_template,
                start_date,
                end_date,
                output_dir,
            )
            files = [
                {
                    "filename": path.name,
                    "download_url": register_download(path),
                    "output_path": str(path),
                }
                for path in result["output_files"]
            ]
            json_response(
                self,
                200,
                {
                    "ok": True,
                    "files": files,
                    "output_dir": str(output_dir),
                    "missing": result["missing"],
                },
            )
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def serve_download(self, path_text: str) -> None:
        key = Path(unquote(path_text.removeprefix("/download/"))).name
        path = DOWNLOAD_FILES.get(key)
        if path is None or not path.exists():
            self.send_error(404, "File not found")
            return
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(path.name)}")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def choose_output_dir(self, parsed) -> None:
        try:
            current = parse_qs(parsed.query).get("current", [""])[0]
            path = choose_output_directory(current)
            json_response(self, 200, {"ok": True, "path": path})
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def log_message(self, format: str, *args) -> None:
        print(f"{self.address_string()} - {format % args}")


class ReusableThreadingHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True


def web_main() -> None:
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    server = ReusableThreadingHTTPServer((HOST, PORT), AppHandler)
    print(f"街道问题汇总表生成 UI: http://{HOST}:{PORT}")
    print("按 Ctrl+C 停止服务。")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        threading.Thread(target=server.server_close).start()


def main() -> None:
    if "--web" in sys.argv:
        web_main()
        return
    sys.modules.setdefault("app", sys.modules[__name__])
    from desktop_app import main as desktop_main

    desktop_main()


if __name__ == "__main__":
    main()
